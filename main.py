
import io
import os
import ast
import inspect
import datetime
import calendar
import shutil
import tkinter as tk
from tkinter import ttk, filedialog, simpledialog, messagebox
from PIL import Image, ImageTk
import matplotlib.pyplot as plt
from math import exp, log
from openpyxl import load_workbook
import numpy as np
import pandas as pd
import statsmodels.api as sm



'''
ACTUARIAL FUNCTIONS
'''

def ln(x):
    return log(x)

mortalityTables = {}


defaultMortalityTable = "1967-70Ultimate.xlsx"
defaultRate = 0.00
defaultRateFreq = 1
defaultMortalityTable2 = "1967-70Ultimate.xlsx"



def readMortalityTable(mTable):
    if(mTable is None or mTable.lower()[-5:] not in (".xlsx", ".xlsm")):
        return None
    tableDir = os.path.dirname(os.path.abspath(__file__))
    tablePath = os.path.join(tableDir, "Mortality Tables", mTable)
    if not os.path.isfile(tablePath):
        return None
    wb = load_workbook(tablePath)
    mt = wb.active
    
    mortalityTable = {}
    for mtRow in mt.iter_rows(min_row = 2, values_only = True):
        x = mtRow[0]
        p = {"ultimate": mtRow[1], "select":[]}
        for i in mtRow[2:]:
            if i is None:
                break
            p["select"].append(i)
        if x is not None and p["ultimate"] is not None:
            mortalityTable[x] = p
    return mortalityTable


'''
Probability to survive for a life currently aged "age"
Here "age" is assumed as the real age of the person, bet selectDuration - year count since selection
Thus, selectionAge = (age - selectDuration), i.e. age = 41 and selectDuration represents probability p[40]+1
If there is no such Select probability or selectDuration < 0 (-1 as default), Ultimate probability is returned
'''
def px(age, mTable = None, selectDuration = -1):
    if mTable is None:
        mTable = defaultMortalityTable
    if mTable not in mortalityTables:
        mortalityTables[mTable] = readMortalityTable(mTable)
    if mortalityTables[mTable] is None:
        return None
    
    if age in mortalityTables[mTable]:
        if(selectDuration < 0):
            return mortalityTables[mTable][age]["ultimate"]
        elif (age - selectDuration) in mortalityTables[mTable]:
            if(selectDuration < len(mortalityTables[mTable][age-selectDuration]["select"])):
                return mortalityTables[mTable][age-selectDuration]["select"][selectDuration]
            else:
                return mortalityTables[mTable][age]["ultimate"]
        else:
            return mortalityTables[mTable][age]["ultimate"]
    else:
        return None
    
    
'''
Death probability at age = x. qx = 1 - px
'''    
def qx(age, mTable = None, selectDuration = -1):
    if age is None:
        return None
        
    pr = px(age, mTable, selectDuration)
    if pr is None:
        return None
    return 1 - pr



#Same as px but for 2 people
def p(age = None, age2 = None, mTable = None, mTable2 = None, selectDuration = -1, selectDuration2 = -1, lastSurvivor = False):
    if age is None:
        return None
    
    if age2 is None:
        return px(age, mTable, selectDuration)
          
    else:
        p1 = px(age, mTable, selectDuration)
        p2 = px(age2, mTable2, selectDuration2)
        if p1 is None or p2 is None:
            return None
        
        if(not lastSurvivor):
            return p1 * p2
        return p1 + p2 - p1*p2
    


def q(age = None, age2 = None, mTable = None, mTable2 = None, selectDuration = -1, selectDuration2 = -1, lastSurvivor = False):
    pr = p(age, age2, mTable, mTable2, selectDuration, selectDuration2, lastSurvivor)
    if pr is None:
        return None
    return 1 - pr



#Probability to survive during m years
def pm(age = None, age2 = None, term = 1, mTable = None, mTable2 = None, lastSurvivor = False,
       selectPeriod = 0, selectPeriod2 = 0):
    if age is None or term is None or age < 0 or term < 0:
        return None
    result = 1
    selectYears = 0
    floatTerm = term - int(term)
    term = int(term)
    
    if age2 is None:
        for i in range(term):
            if(selectPeriod > selectYears):
                pr = p(age = age+i, mTable = mTable, selectDuration = selectYears)
                selectYears += 1
            else:
                pr = p(age = age+i, mTable = mTable)
            if pr is not None:
                result *= pr
            else:
                return None
        if (floatTerm > 0):
            prFloat = p(age = age+term, mTable = mTable, selectDuration = selectYears if selectPeriod > selectYears
                        else -1)
            if prFloat is not None:
                result *= 1 - floatTerm*(1 - prFloat)
            else:
                return None
        return result

    selectYears2 = 0
    for i in range(term):
        if (selectPeriod > selectYears):
            if(selectPeriod2 > selectYears2):
                pr = p(age = age+i, age2 = age2+i, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                            selectDuration = selectYears, selectDuration2 = selectYears2)
                selectYears += 1
                selectYears2 += 1
            else:
                pr= p(age = age+i, age2 = age2+i, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                            selectDuration = selectYears)
                selectYears += 1
        else:
            if(selectPeriod2 > selectYears2):
                pr = p(age = age+i, age2 = age2+i, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                            selectDuration2 = selectYears2)
                selectYears2 += 1
            else:
                pr= p(age = age+i, age2 = age2+i, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor)
        if pr is not None:
            result *= pr
        else: 
            return None
    if (floatTerm > 0):
        prFloat = p(age = age+term, age2 = age2+term, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                    selectDuration = selectYears if selectPeriod > selectYears else -1, 
                    selectDuration2 = selectYears2 if selectPeriod2 > selectYears2 else -1)
        if prFloat is not None:
            result *= 1 - floatTerm*(1 - prFloat)
        else:
            return None
    
    return result
        


def qm(age = None, age2 = None, term = 1, mTable = None, mTable2 = None, lastSurvivor = False,
       selectPeriod = 0, selectPeriod2 = 0):
    pr = pm(age, age2, term, mTable, mTable2, lastSurvivor, selectPeriod, selectPeriod2)
    if pr is None:
        return None
    return 1 - pr




'''
Transforms interest rate to annual i
rate - used interest rate
freq - payment frequency
rType - type of the rate: i, v, d or delta
'''
def toAnnualRate(rate,  freq = 1, rType = "i"):
    try:
        if rate is None:
            return None
        if(rType == "i"):
            if(freq == 1):
                return rate
            return (1 + rate/freq)**freq - 1
        elif(rType == "d"):
            return 1/((1 - rate/freq)**freq) - 1
        elif(rType == "v"):
            return (1/rate)**freq - 1
        elif(rType == "δ"):
            return exp(rate) - 1
        else:
            print("Unexpected error occured! Incorrect rate type")
            return None
    except Exception:
        return None


'''
The following functions calculate i, d, v or delta rate
rate - entered interest rate
fromFreq - frequency of entered interest rate
toFreq - frequency of returned interest rate
rateType - type of the entered rate
'''
def i(rate = None, fromFreq = 1, toFreq = 1, rateType = "i"):
    ann_i = toAnnualRate(rate, fromFreq, rateType)
    if ann_i is None:
        return None
    if(toFreq == 1):
        return ann_i
    return toFreq*((1 + ann_i)**(1/toFreq) - 1)

def d(rate = None, fromFreq = 1, toFreq = 1, rateType = "i"):
    ann_i = toAnnualRate(rate, fromFreq, rateType)
    if ann_i is None:
        return None
    return toFreq*(1 - (1 + ann_i)**(-1/toFreq))

def v(rate = None, fromFreq = 1, toFreq = 1, rateType = "i"):
    ann_i = toAnnualRate(rate, fromFreq, rateType)
    if ann_i is None:
        return None
    return 1/(1 + ann_i)**(1/toFreq)

#toFreq needs only for bugfixing purposes. Doesn't affect calculation results
def delta(rate = None, fromFreq = 1, toFreq = 1, rateType = "i"):
    ann_i = toAnnualRate(rate, fromFreq, rateType)
    if ann_i is None:
        return None
    return ln(1 + ann_i)



#Set defaulr values for functions
def defaults(defer, freq, rate, rateFreq, mTable, mTable2):
    if(defer is None or defer < 0):
        defer = 0
    if(freq is None or freq <= 0):
        freq = 1
    if(rate is None or rate < 0):
        rate = defaultRate
    if(rateFreq is None or rateFreq <= 0):
        rateFreq = defaultRateFreq
    if mTable is None:
        mTable = defaultMortalityTable
    if mTable2 is None:
        mTable2 = defaultMortalityTable
    return defer, freq, rate, rateFreq, mTable, mTable2


def deferAge(age, defer):
    if defer is None:
        defer = 0
    if age is None:
        return None
    return int(age+defer)
        

#Private functions
def annuity(age = None, age2 = None, term = None, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0, isIncreasing = False):
   
    filler, freq, rate, rateFreq, mTable, mTable2 = defaults(0, freq, rate, rateFreq, mTable, mTable2)
    res = 0
    ann_rate = toAnnualRate(rate, rateFreq, rateType)
    if ann_rate is None:
        return None
    if age is None and term is not None:
        if(rate != 0):
            if not isIncreasing:
                return (1 - v(ann_rate)**term)/i(rate, rateFreq, freq, rateType)
            else:
                an = annuity(term = term, freq = freq, rate = rate, rateFreq = rateFreq, rateType = rateType)
                return (1/freq + an - (term + 1/freq) * v(rate = ann_rate)**term)/((1 + ann_rate)**(1/freq) - 1)
        if not isIncreasing:
            return term
        else:
            return ((1 + term)/2)*term
    
    if term is not None:
        summands = int(term*freq)
    elif age is not None:  #Whole life annuity (a_x)
        if mTable not in mortalityTables:
            mortalityTables[mTable] = readMortalityTable(mTable)
        if mortalityTables[mTable] is None:
            return None
        if age2 is None:
            summands = int((max(mortalityTables[mTable].keys()) - age)*freq)
        else:
            if mTable2 not in mortalityTables:
                mortalityTables[mTable2] = readMortalityTable(mTable2)
            if mortalityTables[mTable2] is None:
                return None
            summands = int(min(max(mortalityTables[mTable].keys()) - age, 
                               max(mortalityTables[mTable2].keys()) - age2)*freq)
    else:
        return None

    koef = 1
    for k in range(1, summands+1):
        pr = pm(age, age2, k/freq, mTable, mTable2, lastSurvivor, selectPeriod, selectPeriod2)
        v_rate = v(rate = ann_rate)**(k/freq)
        if isIncreasing:
            koef = int((k-1)/freq + 1)
        if pr is None or v_rate is None:
            return None
        res += (koef/freq) * v_rate * pr
    
    return res



def deferred_annuity(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, 
        rateType = "i", mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0,
        annuityValue = None, isIncreasing = False):
    
    defer, freq, rate, rateFreq, mTable, mTable2 = defaults(defer, freq, rate, rateFreq, mTable, mTable2)
    ann_rate = toAnnualRate(rate, rateFreq, rateType)
    if ann_rate is None:
        return None
    v_rate = v(rate = ann_rate)**(defer)
    if age is None:
        p_defer = 1
    else:
        p_defer = pm(age = age, age2 = age2, term = defer, mTable = mTable, mTable2 = mTable2,
            lastSurvivor = lastSurvivor, selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    newAge = deferAge(age, defer)
    newAge2 = deferAge(age2, defer)
    
    if annuityValue is None:
        annuityValue = annuity(age = newAge, age2 = newAge2, term = term, freq = freq, rate = rate, rateFreq = rateFreq,
            rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod-defer, selectPeriod2 = selectPeriod2-defer, isIncreasing = isIncreasing)
    if v_rate is not None and p_defer is not None and annuityValue is not None:
        return v_rate * p_defer * annuityValue
    else:
        return None



#Actuarial functions for users
def a_imm(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    return deferred_annuity(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
      
  

def a_due(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    defer, freq, rate, rateFreq, mTable, mTable2 = defaults(defer, freq, rate, rateFreq, mTable, mTable2)
    annuityValue = annuity(age = deferAge(age, defer), age2 = deferAge(age2, defer), term = term, freq = freq, 
                    rate = rate, rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, 
                    lastSurvivor = lastSurvivor, selectPeriod = selectPeriod-defer, selectPeriod2 = selectPeriod2-defer)
    v_rate = v(rate = rate, fromFreq = rateFreq, toFreq = 1, rateType = rateType)
    if age is None or age < 0:
        if term is None or term < 0:
            return None
        pr = 1
    elif term is not None and term >= 0:
        pr = pm(age = deferAge(age, defer), age2 = deferAge(age2, defer), term = term, mTable = mTable, mTable2 = mTable2,
                lastSurvivor = lastSurvivor, selectPeriod = selectPeriod-defer, selectPeriod2 = selectPeriod2-defer)
    else:
        pr = 1
    
    if v_rate is None or pr is None or annuityValue is None:
        return None
    if term is None or term < 0:
        annuityValue = annuityValue + 1/freq
    else:
        annuityValue = annuityValue + 1/freq*(1 - v_rate**term * pr)
    return deferred_annuity(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2, annuityValue = annuityValue)



def s_imm(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    if term is None or term < 0: 
        return None
    defer, freq, rate, rateFreq, mTable, mTable2 = defaults(defer, freq, rate, rateFreq, mTable, mTable2)
    a = a_imm(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    i_rate = toAnnualRate(rate, rateFreq, rateType)
    if i_rate is None or a is None:
        return None
    return (1 + i_rate)**(term + defer) * a



def s_due(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    if term is None or term < 0: 
        return None
    defer, freq, rate, rateFreq, mTable, mTable2 = defaults(defer, freq, rate, rateFreq, mTable, mTable2)
    a = a_due(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    i_rate = toAnnualRate(rate, rateFreq, rateType)
    if i_rate is None or a is None:
        return None
    return (1 + i_rate)**(term + defer) * a



def Ia_imm(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    
    return deferred_annuity(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2, isIncreasing= True)




def Ia_due(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    defer, freq, rate, rateFreq, mTable, mTable2 = defaults(defer, freq, rate, rateFreq, mTable, mTable2)
    Ia = Ia_imm(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    an_due = a_due(age = age, age2 = age2, term = term, defer = defer, freq = 1, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    v_rate = v(rate  = rate, fromFreq = rateFreq, toFreq = 1, rateType = rateType)
    if age is None or age < 0:
        if term is None or term < 0:
            return None
        pr = 1
    elif term is not None and term >= 0:
        pr = pm(age = age, age2 = age2, term = term+defer, mTable = mTable, mTable2 = mTable2, 
            lastSurvivor = lastSurvivor, selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    else:
        pr = 1
        
    if Ia is not None and an_due is not None:
        if term is None or term < 0:
            return Ia + an_due
        elif v_rate is None or pr is None:
            return None
        return Ia + an_due/freq - term/freq * v_rate**(term+defer) * pr
    return None



def Is_imm(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    if term is None or term < 0: 
        return None
    Ia = Ia_imm(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    i_rate = toAnnualRate(rate, rateFreq, rateType)
    if i_rate is None or Ia is None:
        return None
    return (1 + i_rate)**(term + defer) * Ia




def Is_due(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    if term is None or term < 0: 
        return None
    Ia = Ia_due(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    i_rate = toAnnualRate(rate, rateFreq, rateType)
    if i_rate is None or Ia is None:
        return None
    return (1 + i_rate)**(term + defer) * Ia




def Da_imm(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    if term is None or term < 0 or term != int(term): 
        return None
    Ia = Ia_imm(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    a = a_imm(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    
    if Ia is None or a is None:
        return None
    return (term + 1)*a - Ia


def Da_due(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    if term is None or term < 0 or term != int(term): 
        return None
    Ia = Ia_due(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    a = a_due(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, 
            rateFreq = rateFreq, rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor, 
            selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    
    if Ia is None or a is None:
        return None
    return (term + 1)*a - Ia



def Ds_imm(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    
    if term is None or term < 0 or term != int(term): 
        return None
    Da = Da_imm(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    i_rate = toAnnualRate(rate, rateFreq, rateType)
    if i_rate is None or Da is None:
        return None
    return (1 + i_rate)**(term + defer) * Da



def Ds_due(age = None, age2 = None, term = None, defer = 0, freq = 1, rate = None, rateFreq = None, rateType = "i",
            mTable = None, mTable2 = None, lastSurvivor = False, selectPeriod = 0, selectPeriod2 = 0):
    
    if term is None or term < 0 or term != int(term): 
        return None
    Da = Da_due(age = age, age2 = age2, term = term, defer = defer, freq = freq, rate = rate, rateFreq = rateFreq,
                rateType = rateType, mTable = mTable, mTable2 = mTable2, lastSurvivor = lastSurvivor,
                selectPeriod = selectPeriod, selectPeriod2 = selectPeriod2)
    i_rate = toAnnualRate(rate, rateFreq, rateType)
    if i_rate is None or Da is None:
        return None
    return (1 + i_rate)**(term + defer) * Da



#Contracts
def A(age = None, defer = 0, rate = None, rateFreq = None, rateType = "i", mTable = None, selectPeriod = 0):
    defer, filler, rate, rateFreq, mTable, filler2 = defaults(defer, None, rate, rateFreq, mTable, None)
    if mTable not in mortalityTables:
        mortalityTables[mTable] = readMortalityTable(mTable)
    if mortalityTables[mTable] is not None and age is not None:
        term = max(mortalityTables[mTable].keys()) - age
    else:
        return None
    return A_term(age = age, term = term, defer = defer, rate = rate, rateFreq = rateFreq, rateType = rateType,
                  mTable = mTable, selectPeriod = selectPeriod)



def A_term(age = None, term = 0, defer = 0, rate = None, rateFreq = None, rateType = "i", mTable = None, selectPeriod = 0):
    defer, filler, rate, rateFreq, mTable, filler2 = defaults(defer, None, rate, rateFreq, mTable, None)
    if age is None or term is None:
        return None
    if defer is None:
        defer = 0

    result = 0
    v_rate = v(rate = rate, fromFreq = rateFreq, toFreq = 1, rateType = rateType)
    if v_rate is None:
        return None
    for i in range(term):
        pr = pm(age = age+defer, term = i, mTable = mTable, selectPeriod = selectPeriod-defer)
        qPr = qx(age = age+defer+i, mTable = mTable, selectDuration = i if selectPeriod - defer > i else -1)
        if pr is None or qPr is None: 
            return None
        result += v_rate**(i+1) * pr * qPr
    
    if defer is not None and defer > 0:
        pr = pm(age = age, term = defer, mTable = mTable, selectPeriod = selectPeriod)
        result = result * pr * v_rate**defer
    
    return result


def A_pEnd(age = None, term = 0, rate = None, rateFreq = None, rateType = "i", mTable = None, selectPeriod = 0):
    defer, filler, rate, rateFreq, mTable, filler2 = defaults(0, None, rate, rateFreq, mTable, None)
    if age is None or term is None or term < 0: 
        return None
    if defer is None:
        defer = 0
    v_rate = v(rate = rate, fromFreq = rateFreq, toFreq = 1, rateType = rateType)
    pr = pm(age = age, term = term, mTable = mTable, selectPeriod = selectPeriod)
    if v_rate is None or pr is None:
        return None
    return v_rate**term * pr
    

def A_end(age = None, term = 0, defer = 0, rate = None, rateFreq = None, rateType = "i", mTable = None, selectPeriod = 0):
    defer, filler, rate, rateFreq, mTable, filler2 = defaults(defer, None, rate, rateFreq, mTable, None)
    if age is None or term is None: 
        return None
    if defer is None:
        defer = 0
    ATerm = A_term(age = age, term = term, defer = defer, rate = rate, rateFreq = rateFreq, rateType = rateType,
                   mTable = mTable, selectPeriod = selectPeriod)
    APEnd = A_pEnd(age = age, term = term+defer, rate = rate, rateFreq = rateFreq, rateType = rateType,
                   mTable = mTable, selectPeriod = selectPeriod)
    if ATerm is None or APEnd is None:
        return None
    return ATerm + APEnd


'''
END OF ACTUARIAL FUNCTION PART
'''

###Constants
tabNames = ["Actuarial Calculator", "Mortality Tables", "Equation Solver", "Mathematical Reserves", "Technical Reserves", "Data"]
labelFont = ("Arial", 15,)
labelFont2 = ("Arial", 20, "bold")
entryFont = ("Arial", 15)


def getMortalityTableNames():
    currentDir = os.path.dirname(os.path.abspath(__file__))
    tableDir = os.path.join(currentDir, "Mortality Tables")
    if not os.path.isdir(tableDir):
        return []

    mortalityTables = []
    for fileName in os.listdir(tableDir):
        fullPath = os.path.join(tableDir, fileName)
        if os.path.isfile(fullPath) and fileName.lower().endswith((".xlsx", ".xlsm", ".xls")):
            mortalityTables.append(fileName)

    return mortalityTables


#For GUI visuals to create readable formulas
def createFormulaImage(formula, master, fontSize, dpi=200):
    figure = plt.figure(figsize = (0.01, 0.01))
    figure.text(0, 0, formula, fontsize = fontSize)
    buf = io.BytesIO()   #Saves to temp bufer instead of SSD
    figure.savefig(buf, format = "png", dpi = dpi, transparent = True, bbox_inches = "tight", pad_inches = 0.03)
    plt.close(figure)

    buf.seek(0)     #Starts img reading from the start
    image = Image.open(buf)
    return ImageTk.PhotoImage(image, master = master)   #Tkinter may bug in case of returning just image



#Create elements for GUI
def createLabel(root, text, x, y, font = labelFont):
    label = tk.Label(root, text = text, font = font, bg = "#eeeeee")
    label.place(x = x, y = y)
    return label



def createEntry(root, x, y, width = 170, height = 30):
    ent = tk.Entry(root, font = entryFont, relief = "solid", bd = 2)
    ent.place(x = x, y = y, width = width, height = height)
    return ent



def createInput(root, name, xPos, yPos, comment, dy = 60):
    createLabel(root, name, xPos, yPos)
    entry = createEntry(root, xPos+270, yPos)
    createLabel(root, comment, xPos+450, yPos)
    yPos += dy
    return entry, yPos



def createLine(root, xPos, yPos, height = 3, width = 600, color = "black", dyUp = 20, dy = 0):
    line = tk.Frame(root, bg = color , height = height, width = width)
    line.place(x = xPos, y = yPos - dyUp)
    return yPos + dy


def createButton(root, text, x, y, width, height, command = None, font = ("Times New Roman", 18, "bold"),
                 bg = "#dddddd", activebg = "#cccccc"):
    button = tk.Button(root, text = text , font = font, bg = bg, activebackground = activebg, command = command)
    button.place(x = x, y = y, width = width, height = height)
    return button


def createRadioButton(root, x, y, text, variable, value, font = ("Arial", 15)):
    rButton = tk.Radiobutton(root, text = text, variable = variable, value = value, font = font)
    rButton.place(x = x, y = y)
    return rButton


def createCombobox(root, text, xPos, yPos, varType, varValue, values, width = 170, height = 30,  dy = 60):
    createLabel(root, text, xPos, yPos)
    if (varType == "str"):
        var = tk.StringVar(value = varValue)
    elif (varType == "int"):
        var = tk.IntVar(value = varValue)
    else:
        return
    box = ttk.Combobox(root, textvariable = var, values = values, state="readonly")
    box.place(x = xPos+270, y = yPos, width = width, height = height)
    yPos += dy
    return box, var, yPos



#Entries for calculator
def createEntryList(root, xPos = 650, yPos = 30, x = True, y = True, n = True, m = True, p = True, MT2 = False):
    entries = {}
    rateVar = tk.StringVar(value = "i")
    rateLetters = ["i", "v", "d", "δ"]
    
    if(x):
        entries["x"], yPos = createInput(root, "Age at Entry (x):", xPos, yPos, "(years)")
    if(y):
        entries["y"], yPos = createInput(root, "Age at Entry 2 (y):", xPos, yPos, "(years)")
    if(n):
        entries["n"], yPos = createInput(root, "Policy Term (n):", xPos, yPos, "(years)")
    if(m):
        entries["m"], yPos = createInput(root, "Defer (m):", xPos, yPos, "(years)")
    if(p):
        entries["p"], yPos = createInput(root, "Payment Frequency (p):", xPos, yPos, "(times / year)")

    yPos = createLine(root, xPos, yPos)

    createLabel(root, "Rate type:", xPos, yPos)
    entries["rateType"] = rateVar
    for i in range(len(rateLetters)):
        rb = tk.Radiobutton(root, text = rateLetters[i], variable = rateVar, value = rateLetters[i], 
                            font = ("Times New Roman", 30, "italic"))
        rb.place(x = xPos+270 + i*65, y = yPos-10)
    yPos += 60

    entries["rate"], yPos = createInput(root, "Rate value:", xPos, yPos, "")
    entries["freq"], yPos = createInput(root, "Conversion frequency:", xPos, yPos, "(times / year)")

    yPos = createLine(root, xPos, yPos)

    entries["mTableBox"], entries["mTable"], yPos = createCombobox(root, "Mortality Table:", xPos, yPos,
                                                                   "str", "", getMortalityTableNames())
    if(y or MT2):
        entries["mTable2Box"], entries["mTable2"], yPos = createCombobox(root, "Mortality Table 2:", xPos, yPos, 
                                                                         "str", "", getMortalityTableNames())
    if(not MT2): 
        entries["select"], yPos = createInput(root, "Select period:", xPos, yPos, "(years)")
    if(y):
        entries["select2"], yPos = createInput(root, "Select period 2:", xPos, yPos, "(years)")
    
    
    entries["fractionalBox"], entries["fractional"], yPos = createCombobox(root, "Fractional inputs:", xPos, yPos,
                                                                           "str", "", ["Floor", "Round", "Ceil"])
    entries["decimal"], yPos = createInput(root, "Precision:", xPos, yPos, "(decimals)")
    
    return entries


def createFormulaList(root, formulas, x0 = 40, y0 = 30, fontsize = 18):
    formulaVar = tk.IntVar(value = 0)
    rootFormulas = []

    for i in range(len(formulas)):
        img = createFormulaImage(formulas[i], root, fontSize = fontsize)
        rootFormulas.append(img)
        rb = tk.Radiobutton(root, variable = formulaVar, value = i, image = img, activebackground = "#e9e9e9")
        rb.place(x = x0, y = y0 + i*110)    
    
    return formulaVar, rootFormulas
    

def createResult(root, xPos = 40, yPos = 640):
    resultVar = tk.StringVar(value = "")
    resultLabel = tk.Label(root, textvariable = resultVar, font = ("Arial", 16, "bold"), bg = "#e9e9e9")
    resultLabel.place(x = xPos, y = yPos)
    return resultVar


def clearFrame(frame):
    for widget in frame.winfo_children():
        widget.destroy()



def getEntryValue(entry, fractional, valueType = "number"):
    value = entry.get()
    if isinstance(value, str):
        value = value.strip()
    if(value == ""):
        return None
    
    if(valueType == "number"):
        try:
            value = float(value)
            if(fractional == "Floor"):
                value = int(value)
            elif(fractional == "Round"):
                value = round(value)
            elif(fractional == "Ceil"):
                if(value != int(value)):
                    value = int(value) + 1
                else:
                    value = int(value)
        except Exception:
            return None
    elif(valueType == "str"):
        value = str(value)
    return value


def getInputs(entries, x = True, y = True, n = True, m = True, p = True, MT2 = False):
    inputs = {}
    fractional = entries["fractional"].get()
    showError = False
    errorCodes = []
    
    if(x):
        inputs["x"] = getEntryValue(entries["x"], fractional, "number")
    if(y):
        inputs["y"] = getEntryValue(entries["y"], fractional, "number")
        if inputs["x"] is None:
            showError = True
            errorCodes.append("Second life (y) cannot be used without first life (x)")
    if(n):
        inputs["n"] = getEntryValue(entries["n"], fractional, "number")
    if(m):
        inputs["m"] = getEntryValue(entries["m"], fractional, "number")
        if(inputs["m"] is not None and inputs["m"] < 0):
            showError = True
            errorCodes.append("Defer (m) should be a non-negative integer")
    if(p):
        inputs["p"] = getEntryValue(entries["p"], fractional, "number")
        if inputs["p"] is None:
            inputs["p"] = 1
        elif(inputs["p"] <= 0):
            showError = True
            errorCodes.append("Payment frequency (p) should be a positive integer")
    
    inputs["rateType"] = getEntryValue(entries["rateType"], fractional, "str")
    inputs["rate"] = getEntryValue(entries["rate"], "None", "number")
    if inputs["rate"] is None:
        inputs["rate"] = 0
    inputs["freq"] = getEntryValue(entries["freq"], fractional, "number")
    if(inputs["freq"] is not None and inputs["freq"] <= 0):
        showError = True
        errorCodes.append("Conversion frequency should be a positive integer")
    
    inputs["mTable"] = getEntryValue(entries["mTable"], fractional, "str")
    if(y or MT2):
        inputs["mTable2"] = getEntryValue(entries["mTable2"], fractional, "str")
    
    if(not MT2):
        inputs["select"] = getEntryValue(entries["select"], fractional, "number")
        if inputs["select"] is None:
            inputs["select"] = 0
        elif(inputs["select"] < 0):
            showError = True
            errorCodes.append("Select period should be a non-negative integer")
        if(y):
            inputs["select2"] = getEntryValue(entries["select2"], fractional, "number")
            if inputs["select2"] is None:
                inputs["select2"] = 0
            elif(inputs["select2"] < 0):
                showError = True
                errorCodes.append("Select period (2) should be a non-negative integer")
    inputs["fractional"] = fractional
    inputs["decimal"] = getEntryValue(entries["decimal"], fractional, "number")
    if(inputs["decimal"] is not None and inputs["decimal"]  < 0):
        showError = True
        errorCodes.append("PDecimal count should be a non-negative integer")
    
    
    return inputs, showError, errorCodes


def setDefaults(inputs):
    global defaultRate, defaultRateFreq, defaultMortalityTable, defaultMortalityTable2
    defaultRate = inputs.get("rate")
    defaultRateFreq = inputs.get("freq")
    defaultMortalityTable = inputs.get("mTable")
    defaultMortalityTable2 = inputs.get("mTable2")

#Section 1 - Single Life

def calcSingleLife(entries, formulaVar, resultVar):
    try:
        inputs, showError, errorCodes = getInputs(entries, y = False)
        setDefaults(inputs)
        if showError:
            print("ERROR")
            print(errorCodes)
        
        formulaIndex = formulaVar.get()
        if(formulaIndex == 0):
            print(inputs)
            result = a_imm(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 1):
            result = a_due(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 2):
            result = s_imm(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 3):
            result = s_due(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        else:
            print("Unknown formula")
        
        dec = inputs["decimal"]
        if dec is None:
            dec = 5
        resultVar.set(f"{round(result, int(dec))}")
        
    except Exception as e:
        print(e)
        resultVar.set(f"Error: {e}")


def createSingleLifeSection(root):
    formulas = [r"${}_{m|}a_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}\ddot{a}_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}s_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}\ddot{s}_{x:\overline{n}|}^{(p)}$"]
    formulaVar, rootFormulas = createFormulaList(root, formulas)
    entries = createEntryList(root, y = False)
    root.tab1S1Img = rootFormulas
    
    resultVar = createResult(root)
    createButton(root, text = "Calculate", x = 40, y = 560, width = 160, height = 60, 
                 command = lambda: calcSingleLife(entries, formulaVar, resultVar))
    
    return formulaVar, entries


#Section 2 - Joint Life

def calcJointLife(entries, formulaVar, resultVar):
    try:
        inputs, showError, errorCodes = getInputs(entries)
        setDefaults(inputs)
        if showError:
            print("ERROR")
            print(errorCodes)
        
        formulaIndex = formulaVar.get()
        if(formulaIndex == 0):
            print(inputs)
            result = a_imm(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = False,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        elif(formulaIndex == 1):
            result = a_due(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = False,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        elif(formulaIndex == 2):
            result = s_imm(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = False,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        elif(formulaIndex == 3):
            result = s_due(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = False,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        else:
            print("Unknown formula")
        
        dec = inputs["decimal"]
        if dec is None:
            dec = 5
        resultVar.set(f"{round(result, int(dec))}")
        
    except Exception as e:
        print(e)
        resultVar.set(f"Error: {e}")

def createJointLifeSection(root):
    formulas = [r"${}_{m|}a_{x y:\overline{n}|}^{(p)}$",
                r"${}_{m|}\ddot{a}_{x y:\overline{n}|}^{(p)}$",
                r"${}_{m|}s_{x y:\overline{n}|}^{(p)}$",
                r"${}_{m|}\ddot{s}_{x y:\overline{n}|}^{(p)}$"]
    formulaVar, rootFormulas = createFormulaList(root, formulas)
    entries = createEntryList(root)
    root.tab1S2Img = rootFormulas
    
    resultVar = createResult(root)
    createButton(root, text = "Calculate", x = 40, y = 560, width = 160, height = 60, 
                 command = lambda: calcJointLife(entries, formulaVar, resultVar))
    
    return formulaVar, entries

#Section 3 - Last Survivor

def calcLastSurvivor(entries, formulaVar, resultVar):
    try:
        inputs, showError, errorCodes = getInputs(entries)
        setDefaults(inputs)
        if showError:
            print("ERROR")
            print(errorCodes)
        
        formulaIndex = formulaVar.get()
        if(formulaIndex == 0):
            print(inputs)
            result = a_imm(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = True,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        elif(formulaIndex == 1):
            result = a_due(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = True,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        elif(formulaIndex == 2):
            result = s_imm(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = True,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        elif(formulaIndex == 3):
            result = s_due(age = inputs.get("x"), age2 = inputs.get("y"),  term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), lastSurvivor = True,
                           selectPeriod = inputs.get("select"), selectPeriod2 = inputs.get("select2"))
        else:
            print("Unknown formula")
        
        dec = inputs["decimal"]
        if dec is None:
            dec = 5
        resultVar.set(f"{round(result, int(dec))}")
        
    except Exception as e:
        print(e)
        resultVar.set(f"Error: {e}")



def createLastSurvivorSection(root):
    formulas = [r"${}_{m|}a_{\overline{x y}:\overline{n}|}^{(p)}$",
                r"${}_{m|}\ddot{a}_{\overline{x y}:\overline{n}|}^{(p)}$",
                r"${}_{m|}s_{\overline{x y}:\overline{n}|}^{(p)}$",
                r"${}_{m|}\ddot{s}_{\overline{x y}:\overline{n}|}^{(p)}$"]
    formulaVar, rootFormulas = createFormulaList(root, formulas)
    entries = createEntryList(root)
    root.tab1S3Img = rootFormulas
    
    resultVar = createResult(root)
    createButton(root, text = "Calculate", x = 40, y = 560, width = 160, height = 60, 
                 command = lambda: calcLastSurvivor(entries, formulaVar, resultVar))
    
    return formulaVar, entries


#Section 4 - Contracts

def calcContracts(entries, formulaVar, resultVar):
    try:
        inputs, showError, errorCodes = getInputs(entries, y = False, p = False)
        setDefaults(inputs)
        if showError:
            print("ERROR")
            print(errorCodes)
        
        formulaIndex = formulaVar.get()
        print(inputs)
        if(formulaIndex == 0):
            result = A(age = inputs.get("x"), defer = inputs.get("m"), rateType = inputs.get("rateType"),
                       selectPeriod = inputs.get("select"))
        elif(formulaIndex == 1):
            result = A_term(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                            rateType = inputs.get("rateType"), selectPeriod = inputs.get("select"))
        elif(formulaIndex == 2):
            result = A_pEnd(age = inputs.get("x"), term = inputs.get("n"), rateType = inputs.get("rateType"),
                             selectPeriod = inputs.get("select"))
        elif(formulaIndex == 3):
            result = A_end(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                            rateType = inputs.get("rateType"), selectPeriod = inputs.get("select"))
        else:
            print("Unknown formula")
        
        dec = inputs["decimal"]
        if dec is None:
            dec = 5
        resultVar.set(f"{round(result, int(dec))}")
        
    except Exception as e:
        print(e)
        resultVar.set(f"Error: {e}")



def createContractsSection(root):
    formulas = [r"${}_{m|}A_{x}$",
                r"${}_{m|}A_{x:\overline{n}|}^{1}$",
                r"${}A_{x:\overline{n}|}^{\,\,\,\,\,\,\,\,1}$",
                r"${}_{m|}A_{x:\overline{n}|}$",]
    formulaVar, rootFormulas = createFormulaList(root, formulas)
    entries = createEntryList(root, y = False, p = False)
    root.tab1S4Img = rootFormulas
    
    resultVar = createResult(root)
    createButton(root, text = "Calculate", x = 40, y = 560, width = 160, height = 60, 
                 command = lambda: calcContracts(entries, formulaVar, resultVar))
    
    return formulaVar, entries

#Section 5 - Increasing annuities

def calcIncreasingAnnuities(entries, formulaVar, resultVar):
    try:
        inputs, showError, errorCodes = getInputs(entries, y = False)
        setDefaults(inputs)
        if showError:
            print("ERROR")
            print(errorCodes)
        
        formulaIndex = formulaVar.get()
        if(formulaIndex == 0):
            print(inputs)
            result = Ia_imm(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 1):
            result = Ia_due(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 2):
            result = Is_imm(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 3):
            result = Is_due(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        else:
            print("Unknown formula")
        
        dec = inputs["decimal"]
        if dec is None:
            dec = 5
        resultVar.set(f"{round(result, int(dec))}")
        
    except Exception as e:
        print(e)
        resultVar.set(f"Error: {e}")   


def createIncreasingAnnuitiesSection(root):
    formulas = [r"${}_{m|}(Ia)_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}(I\ddot{a})_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}(Is)_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}(I\ddot{s})_{x:\overline{n}|}^{(p)}$"]
    formulaVar, rootFormulas = createFormulaList(root, formulas)
    entries = createEntryList(root, y = False)
    root.tab1S5Img = rootFormulas
    
    resultVar = createResult(root)
    createButton(root, text = "Calculate", x = 40, y = 560, width = 160, height = 60, 
                 command = lambda: calcIncreasingAnnuities(entries, formulaVar, resultVar))
    
    return formulaVar, entries

#Section 6 - Decreasing annuities

def calcDecreasingAnnuities(entries, formulaVar, resultVar):
    try:
        inputs, showError, errorCodes = getInputs(entries, y = False)
        setDefaults(inputs)
        if showError:
            print("ERROR")
            print(errorCodes)
        
        formulaIndex = formulaVar.get()
        if(formulaIndex == 0):
            print(inputs)
            result = Da_imm(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 1):
            result = Da_due(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 2):
            result = Ds_imm(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        elif(formulaIndex == 3):
            result = Ds_due(age = inputs.get("x"), term = inputs.get("n"), defer = inputs.get("m"),
                           freq = inputs.get("p"), rateType = inputs.get("rateType"), 
                           selectPeriod = inputs.get("select"))
        else:
            print("Unknown formula")
        
        dec = inputs["decimal"]
        if dec is None:
            dec = 5
        resultVar.set(f"{round(result, int(dec))}")
        
    except Exception as e:
        print(e)
        resultVar.set(f"Error: {e}")   


def createDecreasingAnnuitiesSection(root):
    formulas = [r"${}_{m|}(Da)_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}(D\ddot{a})_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}(Ds)_{x:\overline{n}|}^{(p)}$",
                r"${}_{m|}(D\ddot{s})_{x:\overline{n}|}^{(p)}$"]
    formulaVar, rootFormulas = createFormulaList(root, formulas)
    entries = createEntryList(root, y = False)
    root.tab1S6Img = rootFormulas
    
    resultVar = createResult(root)
    createButton(root, text = "Calculate", x = 40, y = 560, width = 160, height = 60, 
                 command = lambda: calcDecreasingAnnuities(entries, formulaVar, resultVar))
    
    return formulaVar, entries




def showSection(contentFrame, sectionName):
    clearFrame(contentFrame)
    if(sectionName == "Single Life"):
        formulaVar, entries = createSingleLifeSection(contentFrame)
        
    elif(sectionName == "Joint Life"):
        formulaVar, entries = createJointLifeSection(contentFrame)
        
    elif(sectionName == "Last Survivor"):
        formulaVar, entries = createLastSurvivorSection(contentFrame)

    elif(sectionName == "Contracts"):
        formulaVar, entries = createContractsSection(contentFrame)
        
    elif(sectionName == "Increasing annuities"):
        formulaVar, entries = createIncreasingAnnuitiesSection(contentFrame)
        
    elif(sectionName == "Decreasing annuities"):
        formulaVar, entries = createDecreasingAnnuitiesSection(contentFrame)
        
        
        
#######Tab 1 - calculator
def createTab1(root):
#######Sections
    sectionVar = tk.StringVar(value = "Single Life")
    sectionBox = ttk.Combobox(root, textvariable = sectionVar, 
                              values = ["Single Life", "Joint Life", "Last Survivor", "Contracts", 
                                        "Increasing annuities", "Decreasing annuities"],
                              state="readonly")
    sectionBox.place(x = 50, y = 30, width = 180, height = 30)
    contentFrame = tk.Frame(root, bg="#eeeeee")
    contentFrame.place(x=0, y=70, width=1300, height=880)
    
    def onSectionChange(event = None):
        showSection(contentFrame, sectionVar.get())
    
    sectionBox.bind("<<ComboboxSelected>>", onSectionChange)
    showSection(contentFrame, sectionVar.get())


'''
MORTALITY TABLES
'''

#Mortality table functions
MTColnames = ["x", "px", "p[x]"]

def lxToPx(lxFrom, lxTo):
    try:
        return lxTo / lxFrom
    except Exception:
        return None


def getLxFromDx(x, lst):
    if(x is None) or (lst is None) or (len(lst) < x):
        return None
    lx = 0
    for i in range(x, len(lst)):
        lx += lst[i]
    return lx

def dxToPx(dx, lx):
    try:
        return 1 - dx/lx
    except Exception:
        return None


def qxToPx(qx):
    try:
        return 1 - qx
    except Exception:
        return None


def validateMTValues(lst, inputType):
    for i in range(len(lst)):
        if(lst[i] < 0):
            messagebox.showerror("Error", "Input values should not be negative")
            return False
        elif(inputType == 0 and i != len(lst)-1):
            if(lst[i] < lst[i+1]):
                messagebox.showerror("Error", "lx values shoud be descending!")
                return False
        elif(inputType == 2 or inputType == 3):
            if(lst[i] < 0 or lst[i] > 1):
                messagebox.showerror("Error", "Probability is out of [0; 1] range")
                return False
    return True


#Functions for MT creation
def getNextMTColname(col):
    if (col < len(MTColnames)):
        return MTColnames[col]
    return "p[x]+" + str(col - len(MTColnames) + 1)


def selectFile(root, label, doSaveToRoot = True):
    filePath = filedialog.askopenfilename(title = "Select input file", filetypes = [("Excel files", "*.xlsx *.xlsm *.xls")])
    if not filePath:
        return None
    if doSaveToRoot:
        root.selectedInputFile = filePath
        label.config(text = "Input file:" + os.path.basename(filePath))
    
    if not filePath:
        return None
    return filePath



def getMTInputs(selectedInput, tableEntries, HPEntries):
    inputs = {}
    if(selectedInput == 4):
        try:
            for key in HPEntries:
                inputs[key] = float(HPEntries[key].get().strip())
        except Exception as e:
            print(f"{e}")
            print("All Heligman-Pollard parameters should be numbers!")
            return None
    elif(selectedInput in [0, 1, 2, 3]):
        for key in tableEntries:
            inputs[key] = tableEntries[key].get().strip()
        try:
            inputs["firstRowEntry"] = int(inputs["firstRowEntry"])
            inputs["lastRowEntry"] = int(inputs["lastRowEntry"])
        except Exception as e:
            print(f"{e}")
            print("First and last row numbers should be integers!")
            return None
    else:
        return None
    return inputs


def prepareMT(selectedInput, inputs):
    filePath = selectFile(None, None, False)
    doOverwrite = False
    col = 1
    if filePath is None:
        return None
    
    startAge = simpledialog.askinteger("Start age", "Enter Start Age:")
    if startAge is None:
        return None
    if(selectedInput == 4):
        endAge = simpledialog.askinteger("End age", "Enter End Age:")
        if endAge is None:
            return None
        if endAge < startAge:
            messagebox.showerror("Error", "End Age should be greater or equal to Start Age")
            return None
        ageCount = endAge-startAge+1
    else:
        firstRow = inputs["firstRowEntry"]
        lastRow = inputs["lastRowEntry"]
        if lastRow < firstRow:
            messagebox.showerror("Error", "Last Row should be greater or equal to First Row.")
            return None
        ageCount = lastRow-firstRow+1
    
    try:
        wb = load_workbook(filePath)
        mt = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"Could not open Excel file:\n{filePath}")
        print(f"{e}")
        return None

    if mt.cell(row = 1, column = 1).value is None:      #empty cell
        mt.cell(row = 1, column = 1).value = MTColnames[0]
        for i in range(startAge + ageCount):
            if (not doOverwrite) and (mt.cell(row = startAge+i+2, column = 1).value is None):
                mt.cell(row = startAge+i+2, column = 1).value = startAge + i
            else:
                doOverwrite = messagebox.askyesno("Value overwrite exception", "There already is data in cells that script is trying to change. Do you want to overwrite it?")
                if not doOverwrite:
                    return None
    elif mt.cell(row = 1, column = 1).value != MTColnames[0]:
        messagebox.showerror("Wrong file format", f"Incorrect Mortality Table format!\nExpected A1 = {MTColnames[0]}")
        return None
    
    while True:
        currentColname = mt.cell(row = 1, column = col).value
        if currentColname is None:
            break
        expectedColname = getNextMTColname(col - 1)
        if currentColname != expectedColname:
            messagebox.showerror("Wrong file format", f"Incorrect Mortality Table format!\nExpected column name = {expectedColname}, got {currentColname}")
            return None
        col += 1
    newColName = getNextMTColname(col - 1)
    mt.cell(row = 1, column = col).value = newColName
    
    if not doOverwrite:
        for i in range(startAge, startAge + ageCount):
            if(not doOverwrite) and (mt.cell(row = i+2, column = col).value is not None):
                doOverwrite = messagebox.askyesno("Value overwrite exception", "There already is data in cells that script is trying to change. Do you want to overwrite it?")
                if not doOverwrite:
                    return None
                
    return {"filePath": filePath, "workbook": wb, "mortalityTable": mt, "newColumn": col, 
            "newColumnName": newColName, "startAge": startAge, "ageCount": ageCount}



def getColNumber(columnEntry):
    colname = columnEntry.strip().upper()
    if(colname == ""):
        return None
    
    result = 0
    for i in range(len(colname)):
        if not ("A" <= colname[i] <= "Z"):
            return None
        result = result*26 + (ord(colname[i]) - 64)  #64 = ord("A") + 1
    
    return result


def parseExcelCell(cell):
    cellText = cell.strip().upper()
    if cell == "":
        return None
    letters = ""
    digits = ""

    for ch in range(len(cellText)):
        if cellText[ch].isalpha():
            if digits != "":  #digits go before letters like 6A, not A6; or A6B - also wrong
                return None
            letters += cellText[ch]
        elif cellText[ch].isdigit():
            digits += cellText[ch]
        else:
            return None
        
    if(letters == "" or digits == ""):
        return None
    col = 0
    for ch in range(len(letters)):
        col = col*26 + (ord(letters[ch]) - 64)
    row = int(digits)
    if(row <= 0 or col <= 0):
        return None
    return row, col


def getMTValues(root, selectedInput, inputs):
    if(selectedInput == 4):  #Heligman-Polard -> no need to extract data
        return inputs
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select input file first!")
        return None
    
    inputPath = root.selectedInputFile
    columnNumber = getColNumber(inputs["columnEntry"])
    if columnNumber is None:
        messagebox.showerror("Error", "Column name should be a letter: A, B, C, ...,  AA, etc.")
        return None
    firstRow = inputs["firstRowEntry"]
    lastRow = inputs["lastRowEntry"]

    try:
        wb = load_workbook(inputPath, data_only = True)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"Could not open input file:\n{inputPath}")
        print(f"{e}")
        return None

    values = []
    for i in range(firstRow, lastRow + 1):
        value = ws.cell(row= i, column = columnNumber).value
        try:
            value = float(value)
        except Exception:
            messagebox.showerror("Error", f"All selected input values should be numbers.\nInvalid value at row {i}")
            return None
        values.append(value)
    return values


def addMTColumn(root, tableInputVar, tableEntries, HPEntries):
    selectedInput = tableInputVar.get()
    inputs = getMTInputs(selectedInput, tableEntries, HPEntries)
    outputs = []
    if inputs is None:
        return
    MTData = prepareMT(selectedInput, inputs)
    if MTData is None:
        return
    values = getMTValues(root, selectedInput, inputs)
    if values is None:
        return
    print(f" Inputs: {inputs}\n MTData: {MTData}\n Values: {values}\n")
    
    if(selectedInput == 4) or (validateMTValues(values, selectedInput)):
        if(selectedInput == 0):       #lx
            for i in range(len(values)-1):
                outputs.append(lxToPx(values[i], values[i+1]))
        elif(selectedInput == 1):       #dx
            lx = []
            for i in range(len(values)):
                lx.append(getLxFromDx(i, values))
            for i in range(len(lx)-1):
                outputs.append(lxToPx(lx[i], lx[i+1]))
        elif(selectedInput == 2):       #px
            outputs = values.copy()
        elif(selectedInput == 3):       #qx
            for i in range(len(values)):
                outputs.append(qxToPx(values[i]))
        elif(selectedInput == 4):       #HP
            for x in range(MTData["startAge"], MTData["startAge"] + MTData["ageCount"]):
                coef1 = inputs["A"]**((x + inputs["B"])**inputs["C"])
                coef2 = inputs["D"]*exp(-inputs["E"]*(ln(x) - ln(inputs["F"]))**2)
                coef3 = inputs["G"] * inputs["H"]**x
                output = 1/(1 + coef1 + coef2 + coef3)
                if(output < 0) or (output > 1):
                    messagebox.showerror("Error", f"Invalid coefficients entered!\nProbability {output} at age {x} produced")
                    return None
                outputs.append(output)
        else:
            messagebox.showerror("Error", "Invalid input type: {selectedInput}")
            return None
    
    mt = MTData["mortalityTable"]
    for i in range(len(outputs)):
        mt.cell(row = MTData["startAge"] + i + 2, column = MTData["newColumn"]).value = outputs[i]
    MTData["workbook"].save(MTData["filePath"])



def updMTBox(mTableBox, mTableVar, name):
    newValues = getMortalityTableNames()
    mTableBox["values"] = newValues
    mTableVar.set(name)
    return mTableBox, mTableVar



def openMT(mTableVar):
    filename = mTableVar.get().strip()
    if filename == "":
        messagebox.showwarning("Open mortality table", "Please select a mortality table first!")
        return
    currentDir = os.path.dirname(os.path.abspath(__file__))
    tablePath = os.path.join(currentDir, "Mortality Tables", filename)
    if not os.path.isfile(tablePath):           #To check if file was removed / deleted during run
        messagebox.showerror("Open mortality table", "Selected file was not found.")
        return
    try:
        os.startfile(tablePath)
    except Exception as e:
        messagebox.showerror("Open mortality table", f"Could not open file\n\n{e}")



'''
renameMT procedure renames a selected Mortality Table. mTableVar contains table name, mTableBox - box with MT list.
User inputs new file name, if it ends with .xlsx, .xlsm, .xls - use this extension. Else save original extension.
Crashes if 1) No MT selected 2) It was cut/deleted during run 3) User cancel 4) try-except
'''
def renameMT(mTableVar, mTableBox):
    filename = mTableVar.get().strip()
    if filename == "":
        messagebox.showwarning("Rename mortality table", "Please select a mortality table first!")
        return
    currentDir = os.path.dirname(os.path.abspath(__file__))
    filepath = os.path.join(currentDir, "Mortality Tables", filename)
    if not os.path.isfile(filepath):
        messagebox.showerror("Rename mortality table", "File not found.")
        return

    newName = simpledialog.askstring("Rename mortality table", "Enter new file name:")
    if not newName:
        return 
    newName = newName.strip()

    validExt = (".xlsx", ".xlsm", ".xls")
    name, ext = os.path.splitext(newName)   #name - filler
    if ext.lower() not in validExt:
  #splittext[0] - everything before extension, splittext[1] - extension. See: https://www.geeksforgeeks.org/python/python-os-path-splitext-method/
        ext = os.path.splitext(filename)[1]
        newName = newName + ext

    newPath = os.path.join(currentDir, "Mortality Tables", newName)
    try:
        os.rename(filepath, newPath)
        updMTBox(mTableBox, mTableVar, name = newName)
        messagebox.showinfo("Rename mortality table", "File renamed successfully.")
    except Exception as e:
        messagebox.showerror("Rename mortality table", f"Error:\n{e}")



def deleteMT(mTableVar, mTableBox):
    filename = mTableVar.get().strip()
    if filename == "":
        messagebox.showwarning("Delete mortality table", "Please select a mortality table first!")
        return
    currentDir = os.path.dirname(os.path.abspath(__file__))
    tablePath = os.path.join(currentDir, "Mortality Tables", filename)
    if not os.path.isfile(tablePath):
        messagebox.showerror("Delete mortality table", "Selected file was not found.")
        return
    confirm = messagebox.askyesno("Delete mortality table", f"Delete file '{filename}'?")
    if not confirm:
        return
    
    try:
        os.remove(tablePath)
        updMTBox(mTableBox, mTableVar, name = "")
        messagebox.showinfo("Delete mortality table", f"Mortality table {filename} deleted successfully.")
    except Exception as e:
        messagebox.showerror("Delete mortality table", f"Could not delete file.\n\n{e}")



def createTableInputs(root, x = 100, y = 100, dx = 130, dy = 60):
    createLabel(root, "Column:", x, y)
    columnEntry = createEntry(root, x + dx, y)
    
    createLabel(root, "First row:", x, y+dy)
    firstRowEntry = createEntry(root, x + dx, y + dy)
    
    createLabel(root, "Last row:", x, y+2*dy)
    lastRowEntry = createEntry(root, x + dx, y + 2*dy)
    return columnEntry, firstRowEntry, lastRowEntry


#Creates entries for Heligman-Pollard 
def createHPInputs(root, x = 110, y = 70, dx = 180, dy = 80):
    HPEntries = {}
    arguments = ["A", "B", "C", "D", "E", "F", "G", "H"]
    for i in range(len(arguments)):
        createLabel(root, f"{arguments[i]}:", x + i%2 * dx, y + i//2 * dy)
        HPEntries[arguments[i]] = createEntry(root, x + 50 + + i%2 * dx, y + i//2 * dy, width = 100)

    return HPEntries



def createTab2(root):
    titleLabel = createLabel(root, "Create", 30, 35, font = labelFont2)
    tableInputVar, formulaImages = createFormulaList(root, [r"$l_x$", r"$d_x$", r"$p_x$", r"$q_x$", "Use Heligman-Pollard model"], 
                      x0 = 25, y0 = 200, fontsize = 14)
    colAddButton = createButton(root, text = "Add Column", x = 30, y = 730, width = 175, height = 55,
                 command = lambda: addMTColumn(root, tableInputVar, {"columnEntry":columnEntry, 
                                               "firstRowEntry":firstRowEntry, "lastRowEntry":lastRowEntry}, HPEntries))
    
    selectedFileLabel = createLabel(root, "", 160, 140)
    inputButton = createButton(root, text = "Select Input File", x = 280, y = 40, width = 210, height = 60, 
                 command = lambda: selectFile(root, selectedFileLabel))
    
    tableFrame = tk.Frame(root, bg = "#eeeeee")
    tableFrame.place(x = 150, y = 170, width = 450, height = 400)
    columnEntry, firstRowEntry, lastRowEntry = createTableInputs(tableFrame)
    
    HPFrame = tk.Frame(root, bg = "#eeeeee")
    HPFrame.place(x = 150, y = 170, width = 450, height = 400)
    HPEntries = createHPInputs(HPFrame)
    HPFrame.place_forget()
    
    def updMTInputs(*args):
        if(tableInputVar.get() == 4):
            tableFrame.place_forget()
            HPFrame.place(x = 150, y = 170, width = 450, height = 400)
        else:
            HPFrame.place_forget()
            tableFrame.place(x = 150, y = 170, width = 450, height = 400)
    tableInputVar.trace_add("write", updMTInputs)
    
    vertLine = tk.Frame(root, bg = "black", width = 2, height = 650)
    vertLine.place(x = 675, y = 100)
    
    title2Label = createLabel(root, "Mortality Tables", 730, 35, font = labelFont2)
    mTableBox, mTableVar, filler = createCombobox(root, text = "", xPos = 460, yPos = 100, varType = "str",
                                varValue = "", values = getMortalityTableNames(), width = 200, dy = 0)
    openButton = createButton(root, text = "Open", x = 750, y = 200, width = 160, height = 60,
                              command = lambda: openMT(mTableVar))
    renameButton = createButton(root, text = "Rename", x = 750, y = 300, width = 160, height = 60,
                                command = lambda: renameMT(mTableVar, mTableBox))
    deleteButton = createButton(root, text = "Delete", x = 750, y = 400, width = 160, height = 60,
                                command = lambda: deleteMT(mTableVar, mTableBox))
    root.tab2Img = formulaImages
    
    return {
        "labels": (titleLabel, title2Label, selectedFileLabel),
        "buttons": (colAddButton, inputButton, openButton, renameButton, deleteButton),
        "entries": (columnEntry, firstRowEntry, lastRowEntry, HPEntries),
        "boxes": (mTableBox),
        "vars": (tableInputVar, mTableVar)}



'''
EQUATION SOLVER
'''

allowedFunctionNames = {
    "a_imm": a_imm, "a_due": a_due, "s_imm": s_imm, "s_due": s_due,
    "Ia_imm": Ia_imm, "Ia_due": Ia_due, "Is_imm": Is_imm, "Is_due": Is_due,
    "Da_imm": Da_imm, "Da_due": Da_due, "Ds_imm": Ds_imm, "Ds_due": Ds_due,
    "A": A, "A_term": A_term, "A_pEnd": A_pEnd, "A_end": A_end,
    "pm": pm, "qm": qm, "p": p, "q": q,
    "i": i, "d": d, "v": v, "delta": delta, "exp": exp, "ln": ln,
}

'''
This function allows to use only functions and operations that should't be harmful for user
allowedNodeTypes allows to use only expressions (P + 10) binary operaions (+, -, ...) unary operations (-x), 
function call (a_imm()), variable names (P, S), load variables (user can't write P = 10 / del S, etc.), use constants,
arguments in functions (a_due(term = 10)), +, -, *, /, %, +x, -x (negative numbers)                                           
Imports, loops, evals etc. are not allowed to use.
Then equation is parsed to AST Tree - each node of this tree is checked if 1) node has allowed type
2) functions and variables are allowed or local 3) function is called directly and callable function is allowed
'''
def safeEval(expr, localVars, functionNames = None):
    if functionNames is None:
        functionNames = allowedFunctionNames
    allowedNodeTypes = (ast.Expression, ast.BinOp, ast.UnaryOp, ast.Call, ast.Name, ast.Load, ast.Constant, ast.keyword,
                        ast.Add, ast.Sub, ast.Mult, ast.Div, ast.Pow, ast.Mod, ast.USub, ast.UAdd,)

    tree = ast.parse(expr, mode = "eval")
    for node in ast.walk(tree):
        if not isinstance(node, allowedNodeTypes):
            raise ValueError(f"Forbidden expression element: {type(node).__name__}")
        if isinstance(node, ast.Name):
            if node.id not in functionNames and node.id not in localVars:
                raise ValueError(f"Unknown name: {node.id}")
        if isinstance(node, ast.Call):
            if not isinstance(node.func, ast.Name):
                raise ValueError("Only direct function calls are allowed")
            if node.func.id not in functionNames:
                raise ValueError(f"Function is not allowed: {node.func.id}")

    environment = {}
    environment.update(functionNames)
    environment.update(localVars)
    return eval(compile(tree, "<equation>", "eval"), {"__builtins__": {}}, environment)


#Correctly split equation to left and right side
def splitEquation(equation):
    depth = 0
    count = 0
    for i in range(len(equation)):
        if(equation[i] == "("):
            depth += 1
        elif(equation[i] == ")"):
            depth -= 1
        elif(equation[i] == "=") and (depth == 0):
            left = equation[:i]
            right = equation[i+1:]
            count += 1
    if(count == 1):
        return left.strip(), right.strip()
    return None, None


#Equation solving
def solveLinearEquation(equation, variable, constant):
    leftSide, rightSide = splitEquation(equation)
    if leftSide is None or rightSide is None:
        messagebox.showerror("Error", "Equation should contain exactly one '=' outside parentheses!")
        return None
    if(variable == "P"):
        localVars0 = {"P": 0, "S": constant}
        localVars1 = {"P": 1, "S": constant}
    elif(variable == "S"):
        localVars0 = {"S": 0, "P": constant}
        localVars1 = {"S": 1, "P": constant}
    else:
        messagebox.showerror("Error", "Unknown variable!")
        return None

    try:
        f0 = safeEval(leftSide, localVars0) - safeEval(rightSide, localVars0)
        f1 = safeEval(leftSide, localVars1) - safeEval(rightSide, localVars1)
    except Exception as e:
        messagebox.showerror("Error", f"Could not evaluate equation:\n{equation}")
        print(f"{e}")
        return None

    A = f1 - f0
    B = -f0
    if(A == 0):
        messagebox.showerror("Error", "Equation is not solvable: coefficient A in (Ax = B) is zero.")
        return None
    return B / A


def solveActuarialEquation(equationEntry, variableVar, pEntry, sEntry, defaultEntries, resultVar):
    equationText = equationEntry.get().strip()
    if equationText == "":
        messagebox.showerror("Error", "Equation is not entered.")
        return
    equationInputs = getEquationInputs(variableVar, pEntry, sEntry)
    if equationInputs is None:
        return

    defaultsInputs, showError, errorCodes = getInputs(defaultEntries, x = False, y = False, n = False, m = False, 
                                                      p = False, MT2 = True)
    if showError:
        messagebox.showerror("Error", "\n".join(errorCodes))
        return

    setDefaults(defaultsInputs)
    answer = solveLinearEquation(equationText, equationInputs["variable"], equationInputs["constant"])
    if answer is None:
        return
    dec = defaultsInputs.get("decimal")
    if dec is None:
        dec = 5

    resultVar.set(f"{equationInputs['variable']} = {round(answer, int(dec))}")


def updateEquationVariableInputs(variableVar, pEntry, sEntry):
    selected = variableVar.get()
    if selected == "P":
        pEntry.delete(0, tk.END)
        pEntry.config(state = "disabled")
        sEntry.config(state = "normal")
    elif selected == "S":
        sEntry.delete(0, tk.END)
        sEntry.config(state = "disabled")
        pEntry.config(state = "normal")


def getEquationInputs(variableVar, pEntry, sEntry):
    selectedVariable = variableVar.get()

    if selectedVariable == "P":
        try:
            claim = float(sEntry.get().strip())
        except Exception:
            messagebox.showerror("Error", "S (claim sum) should be a number!")
            return None
        return {"variable": "P", "constant": claim}
    elif selectedVariable == "S":
        try:
            premium = float(pEntry.get().strip())
        except Exception:
            messagebox.showerror("Error", "P (premium) should be a number!")
            return None
        return {"variable": "S", "constant": premium}
    else:
        messagebox.showerror("Error", "Unknown variable selected.")
        return None


def createTab3(root):
    createLabel(root, "Equation", 300, 30, font = labelFont2)
    equationEntry = createEntry(root, x = 30, y = 80, width = 620, height = 65)
    
    createLabel(root, "Variable", 45, 220)
    variableVar = tk.StringVar(value = "S")
    pRadio = tk.Radiobutton(root,text = "P", variable = variableVar, value = "P", font = ("Arial", 15, "bold"))
    pRadio.place(x = 40, y = 280)
    pEntry = createEntry(root, x = 170, y = 280, width = 160, height = 32)
    sRadio = tk.Radiobutton(root, text = "S", variable = variableVar, value = "S", font = ("Arial", 15, "bold"))
    sRadio.place(x = 40, y = 340)
    sEntry = createEntry(root, x = 170, y = 340, width = 160, height = 32)

    createLabel(root, "Default values", 910, 30, font=labelFont2)
    entries = createEntryList(root, xPos = 680, yPos = 120, x = False, y = False, n = False, m = False, p = False, MT2 = True)

    variableVar.trace_add("write", lambda *args: updateEquationVariableInputs(variableVar, pEntry, sEntry))
    updateEquationVariableInputs(variableVar, pEntry, sEntry)
    
    resultVar = createResult(root, xPos = 35, yPos = 600)
    createButton(root, text = "Solve", x = 35, y = 520, width = 160, height = 60,
                               command = lambda: solveActuarialEquation( equationEntry, variableVar, pEntry, sEntry,
                                                                        entries, resultVar))
    
'''
MATH RESERVE
'''

EPVFunctionNames = {
    "a_imm", "a_due", "s_imm", "s_due",
    "Ia_imm", "Ia_due", "Is_imm", "Is_due",
    "Da_imm", "Da_due", "Ds_imm", "Ds_due",
    "A", "A_term", "A_pEnd", "A_end"
}


def getReserveInputs(pEntry, sEntry, tEntry):
    try:
        P = float(pEntry.get().strip())
    except Exception:
        messagebox.showerror("Error", "P (premium) should be a number!")
        return None
    try:
        S = float(sEntry.get().strip())
    except Exception:
        messagebox.showerror("Error", "S (claim) should be a number!")
        return None
    try:
        t = int(tEntry.get().strip())
    except Exception:
        messagebox.showerror("Error", "t should be an integer!")
        return None
    if t < 0:
        messagebox.showerror("Error", "t should be non-negative!")
        return None
    return {"P": P, "S": S, "t": t}


#Changes arguments in actuarial functions
def makeReserveFunction(func, t):
    def reserveFunc(*args, **kwargs):
        signature = inspect.signature(func)
        bound = signature.bind_partial(*args, **kwargs)
        params = bound.arguments

        age = params.get("age")
        age2 = params.get("age2")
        term = params.get("term")
        defer = params.get("defer", 0)

        if defer is None:
            defer = 0
        if age is not None:
            params["age"] = age + t
        if age2 is not None:
            params["age2"] = age2 + t
        if "defer" in signature.parameters:
            if t < defer:
                params["defer"] = defer - t
            else:
                params["defer"] = 0
        if "term" in signature.parameters and term is not None:
            if t < defer:
                newTerm = term
            else:
                newTerm = term - (t - defer)
            if newTerm < 0:
                return 0
            params["term"] = newTerm
        return func(**params)
    return reserveFunc



def getReserveFunctionNames(t):
    reserveFunctions = allowedFunctionNames.copy()
    for name in EPVFunctionNames:
        reserveFunctions[name] = makeReserveFunction(allowedFunctionNames[name], t)
    return reserveFunctions



def calculateMathReserve(premiumsEntry, benefitsEntry, expensesEntry, pEntry, sEntry, tEntry, defaultEntries, resultVar):
    premiumsText = premiumsEntry.get().strip()
    benefitsText = benefitsEntry.get().strip()
    expensesText = expensesEntry.get().strip()

    if premiumsText == "":
        premiumsText = "0"
    if benefitsText == "":
        benefitsText = "0"
    if expensesText == "":
        expensesText = "0"

    reserveInputs = getReserveInputs(pEntry, sEntry, tEntry)
    if reserveInputs is None:
        return

    defaultsInputs, showError, errorCodes = getInputs(defaultEntries, x = False, y = False, n = False,
                                                      m = False, p = False, MT2 = True)
    if showError:
        messagebox.showerror("Error", "\n".join(errorCodes))
        return
    setDefaults(defaultsInputs)
    localVars = {"P": reserveInputs["P"], "S": reserveInputs["S"]}
    reserveFunctions = getReserveFunctionNames(reserveInputs["t"])

    try:
        premiums = safeEval(premiumsText, localVars, reserveFunctions)
        benefits = safeEval(benefitsText, localVars, reserveFunctions)
        expenses = safeEval(expensesText, localVars, reserveFunctions)
    except Exception as e:
        messagebox.showerror("Error", "Could not calculate reserve! Please check entered expressions")
        print(f"{e}")
        return
    reserve = benefits + expenses - premiums

    dec = defaultsInputs.get("decimal")
    if dec is None:
        dec = 5
    resultVar.set(f"Reserve = {round(reserve, int(dec))}")



def createTab4(root):
    createLabel(root, "Mathematical Reserves", 220, 30, font=labelFont2)

    createLabel(root, "Premiums flow", 35, 100)
    premiumsEntry = createEntry(root, x = 220, y = 95, width = 430, height = 45)
    createLabel(root, "Benefits flow", 35, 180)
    benefitsEntry = createEntry(root, x = 220, y = 175, width = 430, height = 45)
    createLabel(root, "Expenses flow", 35, 260)
    expensesEntry = createEntry(root, x = 220, y = 255, width = 430, height = 45)

    createLabel(root, "Constants", 45, 350)
    createLabel(root, "t", 40, 410, font=("Arial", 15, "bold"))
    tEntry = createEntry(root, x = 85, y = 410, width = 160, height = 32)
    createLabel(root, "P", 40, 470, font=("Arial", 15, "bold"))
    pEntry = createEntry(root, x = 85, y = 470, width = 160, height = 32)
    createLabel(root, "S", 40, 530, font=("Arial", 15, "bold"))
    sEntry = createEntry(root, x = 85, y = 530, width = 160, height = 32)

    resultVar = createResult(root, xPos = 35, yPos = 700)
    createButton(root, text = "Calculate", x = 35, y = 620, width = 160, height = 60,
                 command = lambda: calculateMathReserve(premiumsEntry, benefitsEntry, expensesEntry, 
                                                        pEntry, sEntry, tEntry, entries, resultVar))

    createLabel(root, "Default values", 910, 30, font=labelFont2)
    entries = createEntryList(root, xPos = 680, yPos = 120, x = False, y = False, n = False, m = False, p = False, MT2 = True)




'''
TECHNICAL RESERVES (IBNR, RBNS, UPR)
'''


policyColnames = ["policy nr", "start date", "end date", "written premium", "coverage end date", "last premium date",
                  "is in force"]
policyObligatoryColnames = ["policy nr", "start date", "end date", "written premium"]
claimColnames = ["claim id", "policy nr", "accident date", "report date", "reported claim amount", "case reserve",
                 "claim status", "upcoming payment count", "upcoming payment amount"]
claimObligatoryColnames = ["claim id", "accident date", "report date", "reported claim amount", "case reserve",
                           "claim status"]

#Mapping for policy files
def getPolicyMapping(filePath, colnames, obligatoryColnames, doValidate = True):
    mapping = {}
    try:
        wb = load_workbook(filePath)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", "Could not open the file!")
        print(f"File {filePath} not found?\n{e}")
        return None, None, None
    
    for i in range(1, ws.max_column + 1):
        colname = ws.cell(row = 1, column = i).value
        if colname is not None:
            colname = colname.strip().lower()
        else:
            continue
        if colname in colnames:
            mapping[colname] = i
    
    if doValidate:
        for colname in obligatoryColnames:
            if colname not in mapping.keys():
                messagebox.showerror("Error", f"Column {colname} is not found!")
                return None, None, None
    return wb, ws, mapping




def createValuationDate(root, x = 800, y = 100, dx = 200):
    createLabel(root, "Valuation Date:", x, y)
    dayEntry = createEntry(root, x + dx, y, width = 50, height = 32)
    createLabel(root, ".", x + dx + 52, y)
    monthEntry = createEntry(root, x + dx + 70, y, width = 50, height = 32)
    createLabel(root, ".", x + dx + 122, y)
    yearEntry = createEntry(root, x + dx + 140, y, width = 100, height = 32)
    createLabel(root, "DD.MM.YYYY", x + dx, y + 45)
    return dayEntry, monthEntry, yearEntry



def getValuationDate(dayEntry, monthEntry, yearEntry):
    try:
        day = int(dayEntry.get().strip())
        month = int(monthEntry.get().strip())
        year = int(yearEntry.get().strip())
        return datetime.date(year, month, day)
    except Exception as e:
        messagebox.showerror("Error", "Invalid valuation date!")
        print(f"{e}")
        return None



def parseDate(date):
    if date is None:
        return None
    if isinstance(date, datetime.datetime):  #Excel format
        return date.date()
    if isinstance(date, datetime.date):    #Excel format
        return date
    if isinstance(date, str):
        date = date.strip().rstrip(".")
        for dateFormat in ["%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"]:
            try:
                return datetime.datetime.strptime(date, dateFormat).date()
            except Exception:
                pass
    return None

    

#UPR calculations
def checkUPRColumn(ws, mapping, outputCol):
    if outputCol is None:
        messagebox.showerror("Error", "Invalid output column! Should be in a letter format A, B, ... , AA, etc.")
        return None
    if outputCol in mapping.values():
        messagebox.showerror("Error", "UPR output column can't be a source data column!")
        return None
    
    for i in range(1, ws.max_row + 1):
        if ws.cell(row = i, column = outputCol).value is not None:
            doOverwrite = messagebox.askyesno("Overwriting warning", "UPR output column is not empty! Do you want to overwrite it?")
            if not doOverwrite:
                return None
            else:
                break
    return outputCol


def getUPRValues(ws, mapping, row):
    policyNr = ws.cell(row = row, column = mapping["policy nr"]).value
    startDate = ws.cell(row = row, column = mapping["start date"]).value
    endDate = ws.cell(row = row, column =  mapping["end date"]).value
    writtenPremium = ws.cell(row = row, column = mapping["written premium"]).value
    if "coverage end date" in mapping:
        coverageEndDate = ws.cell(row = row, column = mapping["coverage end date"]).value
    else:
        coverageEndDate = None
    if "last premium date" in mapping:
        lastPremiumDate = ws.cell(row = row, column = mapping["last premium date"]).value
    else:
        lastPremiumDate = None
    if "is in force" in mapping:
        isInForce = ws.cell(row = row, column = mapping["is in force"]).value
    else:
        isInForce = None
    if(isInForce != False):
        isInForce = True
    try:
        writtenPremium = float(writtenPremium)
    except Exception:
        messagebox.showerror("Error", f"Written premium should be a number.\nInvalid row: {row}")
        return None
    
    return {"policy nr": policyNr, "start date": startDate, "end date": endDate, "written premium": writtenPremium,
            "coverage end date": coverageEndDate, "last premium date": lastPremiumDate, "is in force": isInForce}
        

def calcUPR(inputs, valuationDate):
    startDate = parseDate(inputs["start date"])
    endDate = parseDate(inputs["end date"])
    if not inputs["is in force"]:
        if(valuationDate >= startDate):
            return 0
        else:
            return inputs["written premium"]
    else:
        coverageEndDate = parseDate(inputs["coverage end date"])
        lastPremiumDate = parseDate(inputs["last premium date"])
        if coverageEndDate is not None:
            endDate = min(endDate, coverageEndDate)
        if lastPremiumDate is not None:
            startDate = max(startDate, lastPremiumDate)
        coveredPeriod = max((endDate - valuationDate).days, 0)
        fullPeriod = (endDate - startDate).days
    return min(inputs["written premium"] * coveredPeriod / fullPeriod, inputs["written premium"])
    
    

def UPR(root, dayEntry, monthEntry, yearEntry, exportVar, exportColumnEntry, resultVar):
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select source file first!")
        return None
    filePath = root.selectedInputFile
    valuationDate = getValuationDate(dayEntry, monthEntry, yearEntry)
    wb, ws, mapping = getPolicyMapping(filePath, policyColnames, policyObligatoryColnames)
    if valuationDate is None:
        return None     #Error message already sent 
    if mapping is None:
        return None     #Error message already sent
    
    outputCol = None
    if exportVar.get():
        outputCol = exportColumnEntry.get().strip()
        if(outputCol == ""):
            messagebox.showerror("Error", "Please enter UPR output column or disable output")
            return None
        outputCol = checkUPRColumn(ws, mapping, getColNumber(outputCol))
    
    totalUPR = 0
    for row in range(2, ws.max_row+1):
        UPRValues = getUPRValues(ws, mapping, row)
        upr = calcUPR(UPRValues, valuationDate)
        totalUPR += upr
        if outputCol is not None:
            ws.cell(row = row, column = outputCol).value = upr
    
    if outputCol is not None:
        ws.cell(row = 1, column = outputCol).value = "UPR"
        try:
            wb.save(filePath)
        except Exception as e:
            messagebox.showerror("Error", "Error while saving the file!")
            print(f"{e}")
            return None
    resultVar.set(f"Total UPR = {round(totalUPR, 2)}")



#IBNR functions
def shiftMonths(date, months):
    totalMonth = date.month + months-1    #-1 because of //12 and %12 in next rows
    year = date.year + totalMonth//12
    month = totalMonth%12 + 1
    day = date.day
    lastDay = calendar.monthrange(year, month)[1]
    day = min(day, lastDay)
    return datetime.date(year, month, day)


def shiftDate(date, periodCount, periodLength):
    if(periodLength == "Month"):
        return shiftMonths(date, periodCount)
    elif(periodLength == "Quarter"):
        return shiftMonths(date, periodCount*3)
    elif(periodLength == "Year"):
        return shiftMonths(date, periodCount*12)
    else:
        return None


def getPeriodIndex(date, valuationDate, periodLength, periodCount):
    if(date is None) or (date > valuationDate):
        return None
    for i in range(periodCount):
        periodStart = shiftDate(valuationDate, -(i+1), periodLength)
        periodEnd = shiftDate(valuationDate, -i, periodLength)
        if(periodStart < date <= periodEnd):
            return i
    return None


def createEmptyTriangle(n):
    triangle = []
    for i in range(n):
        triangle.append([])
        for j in range(n):
            triangle[i].append("")
    return triangle
    

def isCumulativeTriangle(triangle):
    for i in range(len(triangle)):
        for j in range(len(triangle[i]) - 1):
            if(triangle[i][j+1] != "" and triangle[i][j + 1] < triangle[i][j]):
                return False
    return True



def cumulativeToIncremental(triangle):
    incTriangle = []
    for i in range(len(triangle)):
        row = []
        for j in range(len(triangle[i])):
            if(j == 0):
                row.append(triangle[i][j])
            elif(triangle[i][j] != ""):
                row.append(triangle[i][j] - triangle[i][j-1])
            else:
                row.append("")
        incTriangle.append(row)
    return incTriangle



def incrementalToCumulative(triangle):
    cumTriangle = []
    for i in range(len(triangle)):
        cum = triangle[i][0]
        cumTriangle.append([cum])
        for j in range(1, len(triangle[i])):
            if(triangle[i][j] != ""):
                cum += triangle[i][j]
                cumTriangle[i].append(cum)
            else:
                cumTriangle[i].append("")
    return cumTriangle
    

def switchIBNRTriangleType(dataTypeVar, dataTypeLabel, triangleFrame, triangleData):
    triangle = triangleData.get("triangle")
    if triangle is None or triangle == []:
        messagebox.showerror("Error", "Please import triangle data!")
        return None
    if dataTypeVar.get() == "Incremental":
        triangle = incrementalToCumulative(triangle)
        dataTypeVar.set("Cumulative")
        dataTypeLabel.config(text = "Cumulative")
    else:
        triangle = cumulativeToIncremental(triangle)
        dataTypeVar.set("Incremental")
        dataTypeLabel.config(text="Incremental")
    triangleData["triangle"] = triangle
    clearFrame(triangleFrame)
    createTrianglePreview(triangleFrame, triangle=triangle)

    

def getPeriodCount(periodCountEntry):
    try:
        periodCount = int(periodCountEntry.get().strip())
    except Exception:
        messagebox.showerror("Error", "Period count should be an integer!")
        return None
    if periodCount <= 0:
        messagebox.showerror("Error", "Period count should be positive!")
        return None
    return periodCount



#Reads loss triangle from excel (in triagle form)
def readTriangleExcel(root, topLeftEntry, periodCountEntry):
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select source file first!")
        return None
    filePath = root.selectedInputFile
    topLeft = parseExcelCell(topLeftEntry.get())
    if topLeft is None:
        messagebox.showerror("Error", "Top left cell is invalid!")
        return None
    periodCount = getPeriodCount(periodCountEntry)
    if periodCount is None:
        return None

    startRow, startCol = topLeft
    triangle = []
    try:
        wb = load_workbook(filePath, data_only=True)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"Could not open Excel file:\n{filePath}")
        print(f"{e}")
        return None

    for i in range(periodCount):
        row = []
        for j in range(periodCount-i):
            value = ws.cell(row = startRow+i, column = startCol+j).value
            try:
                value = float(value)
            except Exception as e:
                messagebox.showerror("Error",
                        f"Could not read triangle data!\nInvalid value at row {startRow+i}, column {startCol+j}.")
                print(value, f"{e}")
                return None
            row.append(value)
        row += [""]*i
        triangle.append(row)
    print(triangle)
    return triangle


#Creates loss triangle from data file
def readTriangleFile(root, periodCountEntry, periodLengthVar, dayEntry, monthEntry, yearEntry):
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select source file first!")
        return None
    filePath = root.selectedInputFile
    valuationDate = getValuationDate(dayEntry, monthEntry, yearEntry)
    if valuationDate is None:
        return None    #Error already sent
    periodCount = getPeriodCount(periodCountEntry)
    if periodCount is None:
        return None    #Error already sent
    periodLength = periodLengthVar.get()

    wb, ws, mapping = getPolicyMapping(filePath, claimColnames, claimObligatoryColnames)
    if mapping is None:     #then wb and ws are None too
        return None
    
    triangle = createEmptyTriangle(periodCount)
    skippedRows = 0
    for row in range(2, ws.max_row + 1):
        accidentDate = ws.cell(row = row, column = mapping["accident date"]).value
        accidentDate = parseDate(accidentDate)
        reportDate = ws.cell(row=row, column = mapping["report date"]).value
        reportDate = parseDate(reportDate)
        reportedClaimAmount = ws.cell(row = row, column = mapping["reported claim amount"]).value
        claimStatus = ws.cell(row = row, column = mapping["claim status"]).value
        if(accidentDate is None) or (reportDate is None) or (reportedClaimAmount is None) or (claimStatus is None):
            skippedRows += 1
            continue
        claimStatus = str(claimStatus).strip().lower()
        if(accidentDate > valuationDate) or (reportDate > valuationDate) or (claimStatus == "rejected"):
            continue
        if(reportDate < accidentDate):
            skippedRows += 1
            continue
        
        try:
            reportedClaimAmount = float(reportedClaimAmount)
        except Exception:
            messagebox.showerror("Error", f"Reported claim amount should be a number!\nInvalid row: {row}")
            #continue
            return None 
        accidentIndex = getPeriodIndex(accidentDate, valuationDate, periodLength, periodCount)
        reportedIndex = getPeriodIndex(reportDate, valuationDate, periodLength, periodCount)
        if(accidentIndex is None) or (reportedIndex is None):
            continue
        
        developmentIndex = accidentIndex - reportedIndex
        originIndex = periodCount - accidentIndex - 1
        if(developmentIndex < 0):
            skippedRows += 1
            continue
        if(originIndex < 0) or (originIndex >= periodCount) or (developmentIndex > periodCount-originIndex-1):
            continue
        if(triangle[originIndex][developmentIndex] != ""):
            triangle[originIndex][developmentIndex] += reportedClaimAmount
        else:
            triangle[originIndex][developmentIndex] = reportedClaimAmount
            
    if skippedRows > 0:
        messagebox.showwarning("Warning", f"{skippedRows} rows were skipped! (invalid dates or invalid periods)")
    return triangle



def importIBNRTriangleData(root, dataSourceVar, periodLengthVar, dataTypeVar, topLeftEntry, periodCountEntry,
                           dayEntry, monthEntry,yearEntry ,dataTypeLabel, triangleFrame, triangleData):
    if dataSourceVar.get() == "Triangle":
        triangle = readTriangleExcel(root, topLeftEntry, periodCountEntry)
        if triangle is None:
            clearFrame(triangleFrame)
            triangleData["triangle"] = []
            createTrianglePreview(triangleFrame, triangle = empty5x5)
            return
        isCumulative = messagebox.askyesno("Data type", "Are imported triangle data cumulative?")
        if isCumulative:
            if isCumulativeTriangle(triangle):
                dataTypeVar.set("Cumulative")
                dataTypeLabel.config(text = "Cumulative")
            else:
                messagebox.showwarning("Warning", "Selected data cannot be cumulative: values decrease inside at least one row.\n"
                                       "Data will be treated as incremental.")
                dataTypeVar.set("Incremental")
                dataTypeLabel.config(text = "Incremental")
        else:
            dataTypeVar.set("Incremental")
            dataTypeLabel.config(text = "Incremental")
    else:
        triangle = readTriangleFile(root, periodCountEntry, periodLengthVar, dayEntry, monthEntry, yearEntry)
        if triangle is None:
            clearFrame(triangleFrame)
            triangleData["triangle"] = []
            createTrianglePreview(triangleFrame, triangle = empty5x5)
            return
        dataTypeVar.set("Incremental")
        dataTypeLabel.config(text = "Incremental")
    triangleData["triangle"] = triangle
    clearFrame(triangleFrame)
    createTrianglePreview(triangleFrame, triangle = triangle)



empty5x5 = [["","","","",""], ["","","","",""], ["","","","",""], ["","","","",""], ["","","","",""]]
def createTrianglePreview(root, triangle = empty5x5, x = 0, y = 0, width = 68, height = 30):
    #Header
    n = len(triangle)
    cells = createEmptyTriangle(n)
    for i in range(n):
        cell = tk.Label(root, text = str(i+1), font = ("Arial", 14, "bold"), bg = "#f5f5f5", relief = "solid", bd = 2)
        cell.place(x = x + (i+1)*width, y = y, width = width, height = height)
    
    #Rows
    for i in range(n):
        cell = tk.Label(root, text = str(i+1), font = ("Arial", 14, "bold"), bg = "#f5f5f5", relief = "solid", bd = 2)
        cell.place(x = x, y = y + (i+1)*height, width = width, height = height)
        for j in range(n):
            color = "#cccccc"
            if(j <= n - i - 1):
                color = "#aaaaaa"
            try:
                value = round(triangle[i][j], 2)
            except:
                value = triangle[i][j]
            cell = tk.Label(root, text = value, font = ("Arial", 9), bg = color, relief = "solid", bd = 2)
            cell.place(x = x + (j+1)*width, y = y + (i+1)*height, width = width, height = height)
            cells[i][j] = cell
    return cells



def showCLInputs(methodFrame):
    clearFrame(methodFrame)
    methodFrame.entries = {}



def showBFInputs(methodFrame):
    clearFrame(methodFrame)
    createLabel(methodFrame, "Premium Sum:", 0, 10)
    premiumSumEntry = createEntry(methodFrame, 135, 10, width = 100)
    createLabel(methodFrame, "ELR:", 0, 60)
    ELREntry = createEntry(methodFrame, 135, 60, width = 100)
    createLabel(methodFrame, "z:", 0, 110)
    zEntry = createEntry(methodFrame, 135, 110, width = 100)
    estimateZVar = tk.BooleanVar(value = False)

    def updateZState(*args):
        if estimateZVar.get():
            zEntry.delete(0, tk.END)
            zEntry.config(state = "disabled")
        else:
            zEntry.config(state = "normal")

    estimateZBox = tk.Checkbutton(methodFrame, text = "Estimate z using CL", variable = estimateZVar,
                                  font = ("Arial", 12), command = updateZState)
    estimateZBox.place(x = 0, y = 155)
    updateZState()
    methodFrame.entries = {"premiumSum": premiumSumEntry, "ELR": ELREntry, "z": zEntry, "estimateZ": estimateZVar}


    
def showGLMInputs(methodFrame):
    clearFrame(methodFrame)
    createLabel(methodFrame, "1", 85, 0, font = ("Arial", 12, "bold"))
    createLabel(methodFrame, "Geom", 140, 0, font = ("Arial", 12, "bold"))
    createLabel(methodFrame, "Param", 215, 0, font = ("Arial", 12, "bold"))

    alphaVar = tk.StringVar(value = "Param")
    betaVar = tk.StringVar(value = "Param")
    gammaVar = tk.StringVar(value = "1")
    rows = [("α", alphaVar, 40), ("β", betaVar, 95), ("γ", gammaVar, 150)]
    
    for label, var, y in rows:
        createLabel(methodFrame, label, 0, y, font=("Times New Roman", 28, "italic"))
        createRadioButton(methodFrame, 85, y + 5, "", var, "1")
        createRadioButton(methodFrame, 145, y + 5, "", var, "Geom")
        createRadioButton(methodFrame, 220, y + 5, "", var, "Param")
    methodFrame.entries = {"alphaMode": alphaVar, "betaMode": betaVar, "gammaMode": gammaVar}



#Calculation functions
def chainLadder(triangleData, getFactorsOnly = False):
    triangle = triangleData["triangle"]
    f = []
    for i in range(len(triangle)-1):
        sumCurrent = 0
        sumNext = 0
        for j in range(len(triangle)-i-1):
            if(triangle[j][i] != "") and (triangle[j][i+1] != ""):
                sumCurrent += triangle[j][i]
                sumNext += triangle[j][i+1]
        try:
            f.append(sumNext/sumCurrent)
        except:
            messagebox.showerror("Error", f"Cannot calculate development factor for period {j + 1}.\nDivision by 0 error!")
            return None
    triangleData["f"] = f
    if getFactorsOnly:
        return
    for i in range(1, len(triangle)):
        for j in range(len(triangle)-i, len(triangle)):
            triangle[i][j] = triangle[i][j-1] * f[j-1]
    IBNR = 0
    for i in range(len(triangle)):
        IBNR += triangle[i][len(triangle)-1] - triangle[i][len(triangle)-1-i]
    triangleData["triangle"] = triangle
    triangleData["f"] = f
    triangleData["method"] = "CL"
    return IBNR


def parseNumberList(text, n, fieldName):
    text = text.strip()
    if text == "":
        messagebox.showerror("Error", f"{fieldName} is empty!")
        return None
    values = [x.strip() for x in text.split(",")]
    try:
        values = [float(x) for x in values]
    except Exception:
        messagebox.showerror("Error", f"{fieldName} should contain numbers separated by commas.")
        return None
    if(len(values) == 1):
        return values*n
    if(len(values) != n):
        messagebox.showerror("Error", f"{fieldName} has incalid value count!\nEnter either one value, or matching period count")
        return None
    return values


#Bornhuetter-Ferguson
def estimateZ(f):
    n = len(f)+1
    z = []
    for i in range(n):
        res = 1
        for j in range(n-i-1, n-1):
            res = res / f[j]
        z.append(res)
    return z


def bornhuetterFerguson(triangleData, premiums, ELR, z):
    triangle = triangleData["triangle"]
    periodIBNR = []
    for i in range(len(triangle)):
        periodIBNR.append(premiums[i] * ELR[i] * (1 - z[i]))
    IBNR = sum(periodIBNR)
    triangleData["method"] = "BF"
    triangleData["periodIBNR"] = periodIBNR
    triangleData["premiums"] = premiums
    triangleData["ELR"] = ELR
    triangleData["z"] = z
    return IBNR



#GLM
def triangleToGLMData(triangle):
    rows = []
    for i in range(len(triangle)):
        for j in range(len(triangle)):
            if triangle[i][j] != "":
                k = i + j
                rows.append({"i": i, "j": j, "k": k, "value": float(triangle[i][j])})
    return pd.DataFrame(rows)



def buildGLMDesignMatrix(df, n, alphaMode, betaMode, gammaMode):
    X = pd.DataFrame(index = df.index)
    X["const"] = 1
    
    def addEffect(mode, prefix, sourceCol, maxLevel):
        if(mode == "1"):
            return         #log(1) = 0
        if(mode == "Geom"):
            X[prefix + "_geom"] = df[sourceCol].astype(float)
            return
        if(mode == "Param"):
            for level in range(1, maxLevel+1):
                X[prefix + "_" + str(level)] = (df[sourceCol] == level).astype(float)
                
    addEffect(alphaMode, "alpha", "i", n-1)
    addEffect(betaMode, "beta", "j", n-1)
    addEffect(gammaMode, "gamma", "k", 2*n - 2)
    return X


def extractGLMFactors(result, n, alphaMode, betaMode, gammaMode):
    params = result.params
    mu = exp(params.get("const", 0))
    
    def getEffectValues(mode, prefix, maxLevel):
        if(mode == "1"):
            return [1 for i in range(maxLevel + 1)]
        if(mode == "Geom"):
            base = exp(params.get(prefix + "_geom", 0))
            return [base**level for level in range(maxLevel + 1)]
        if(mode == "Param"):
            values = [1]
            for level in range(1, maxLevel + 1):
                values.append(exp(params.get(f"{prefix}_{level}", 0)))
            return values
        return [1 for i in range(maxLevel + 1)]
    
    alpha = getEffectValues(alphaMode, "alpha", n - 1)
    beta = getEffectValues(betaMode, "beta", n - 1)
    gamma = getEffectValues(gammaMode, "gamma", 2*n - 2)
    alpha = [mu * x for x in alpha]
    sumBeta = sum(beta)
    beta = [b/sumBeta for b in beta]
    alpha = [a*sumBeta for a in alpha]
    return alpha, beta, gamma



def GLM(triangleData, alphaMode, betaMode, gammaMode):
    triangle = triangleData["triangle"]
    n = len(triangle)
    df = triangleToGLMData(triangle)
    if df.empty:
        messagebox.showerror("Error", "Triangle is empty!")
        return None
    y = df["value"].astype(float)
    X = buildGLMDesignMatrix(df, n, alphaMode, betaMode, gammaMode)
    print(y)
    print(X)
    rank = np.linalg.matrix_rank(X.to_numpy())
    if(rank < X.shape[1]):
        continueCalc = messagebox.askyesno("Warning", "GLM model is not identifiable! Results may be unstable or incorrect.\n"
            "Do you want to continue?")
        if not continueCalc:
            return None

    try:
        model = sm.GLM(y, X, family = sm.families.Poisson())
        result = model.fit()
        alphaValues, betaValues, gammaValues = extractGLMFactors(result, n, alphaMode, betaMode, gammaMode)
    except Exception as e:
        messagebox.showerror("Error", "Could not fit GLM model!")
        print(f"{e}")
        return None

    fullTriangle = createEmptyTriangle(n)
    IBNR = 0
    for i in range(n):
        for j in range(n):
            if(j <= n-i-1):
                fullTriangle[i][j] = triangle[i][j]
            else:
                predDf = pd.DataFrame([{"i": i, "j": j, "k": i + j, "value": 0}])
                Xpred = buildGLMDesignMatrix(predDf, n, alphaMode, betaMode, gammaMode)
                Xpred = Xpred.reindex(columns = X.columns, fill_value = 0)
                pred = float(result.predict(Xpred).iloc[0])
                fullTriangle[i][j] = pred
                IBNR += pred

    triangleData["triangle"] = fullTriangle
    triangleData["method"] = "GLM"
    triangleData["deviance"] = result.deviance
    triangleData["df"] = result.df_resid
    triangleData["glm_modes"] = {"alpha": alphaMode, "beta": betaMode, "gamma": gammaMode}
    triangleData["alpha"] = alphaValues
    triangleData["beta"] = betaValues
    triangleData["gamma"] = gammaValues
    return IBNR



#Main IBNR calc function
def calculateIBNR(methodVar, dataTypeVar, dataTypeLabel, triangleFrame, methodFrame, triangleData, resultVar, GLMInfoVar):
    triangle = triangleData["triangle"]
    if(triangle is None) or (triangle == []):
        messagebox.showerror("Error", "Please import triangle data first!")
        return
    if(methodVar.get() == "CL"):
        if(dataTypeVar.get() == "Incremental"):
            triangleData["triangle"] = incrementalToCumulative(triangle)
            dataTypeVar.set("Cumulative")
            dataTypeLabel.config(text = "Cumulative")
        IBNR = chainLadder(triangleData)
        GLMInfoVar.set("")
        if IBNR is None:
            return
        
    elif methodVar.get() == "BF":
        triangle = triangleData.get("triangle")
        if dataTypeVar.get() == "Incremental":
            triangle = incrementalToCumulative(triangle)
            triangleData["triangle"] = triangle
            dataTypeVar.set("Cumulative")
            dataTypeLabel.config(text="Cumulative")   
        premiums = parseNumberList(methodFrame.entries["premiumSum"].get(), len(triangle), "Premium Sum")
        if premiums is None:
            return None
        ELR = parseNumberList(methodFrame.entries["ELR"].get(), len(triangle), "ELR")
        if ELR is None:
            return
        if methodFrame.entries["estimateZ"].get():
            chainLadder(triangleData, getFactorsOnly = True)
            f = triangleData["f"]
            if f is None:
                messagebox.showerror("Error", "Could not estimate z using Chain Ladder factors.")
                return None
            zValues = estimateZ(f)
        else:
            zValues = parseNumberList(methodFrame.entries["z"].get(), len(triangle), "z")
            if zValues is None:
                return
        GLMInfoVar.set("")
        IBNR = bornhuetterFerguson(triangleData, premiums, ELR, zValues)

    elif methodVar.get() == "GLM":
        if dataTypeVar.get() == "Cumulative":
            triangleData["triangle"] = cumulativeToIncremental(triangle)
            dataTypeVar.set("Incremental")
            dataTypeLabel.config(text="Incremental")
        alphaMode = methodFrame.entries["alphaMode"].get()
        betaMode = methodFrame.entries["betaMode"].get()
        gammaMode = methodFrame.entries["gammaMode"].get()
        IBNR = GLM(triangleData, alphaMode, betaMode, gammaMode)
        GLMInfoVar.set(f"Deviance = {round(triangleData['deviance'], 2)}\n" f"Df = {triangleData['df']}")
        if IBNR is None:
            return
    else:
        messagebox.showerror("Error", "Unknown IBNR method.")
        
    clearFrame(triangleFrame)
    createTrianglePreview(triangleFrame, triangle = triangleData["triangle"])
    triangleData["IBNR"] = IBNR
    resultVar.set(f"IBNR = {round(IBNR, 2)}")



def isExportAreaEmpty(ws, startRow, startCol, rowCount, colCount):
    for i in range(startRow, startRow+rowCount):
        for j in range(startCol, startCol+colCount):
            if ws.cell(row = i, column = j).value is not None:
                return False
    return True


def writeRow(ws, row, startCol, values):
    for i in range(len(values)):
        ws.cell(row = row, column = startCol+i).value = values[i]
        
        
def exportIBNR(triangleData):
    triangle = triangleData.get("triangle")
    method = triangleData.get("method")
    if(triangle is None) or (triangle == []) or (method not in ["CL", "BF", "GLM"]):
        messagebox.showerror("Error", "Nothing to export. Please calculate IBNR first!")
        return None
    filePath = filedialog.askopenfilename(title = "Select Excel file for export", filetypes=[("Excel files", "*.xlsx *.xlsm")])
    if not filePath:
        return None
    topLeft = simpledialog.askstring("Export IBNR", "Enter top left cell (e.g. A6, AB11, ...):")
    if not topLeft:
        return None
    startRow, startCol = parseExcelCell(topLeft)
    if(startRow is None) or (startCol is None):
        messagebox.showerror("Error", "Invalid top left cell!")
        return None
    n = len(triangle)
    colCount = n + 1
    if(method == "CL"):
        rowCount = n + 5
    elif(method == "BF"):
        rowCount = n + 7
    elif(method == "GLM"):
        rowCount = n + 9
        colCount = 2*n   #Because of gammas

    try:
        wb = load_workbook(filePath, keep_vba = filePath.lower().endswith(".xlsm"))
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"Could not open Excel file:\n{e}")
        return None
    if not isExportAreaEmpty(ws, startRow, startCol, rowCount, colCount):
        confirm = messagebox.askyesno("Overwrite warning", "Selected export area is not empty.\nDo you want to overwrite it?")
        if not confirm:
            return None
        
    ws.cell(row = startRow, column = startCol).value = ""
    for i in range(1, n+1):
        ws.cell(row = startRow, column = startCol+i).value = i
    for i in range(1, n+1):
        ws.cell(row = startRow+i, column = startCol).value = i
        for j in range(1, n+1):
            ws.cell(row = startRow+i, column=startCol+j).value = triangle[i-1][j-1]
    infoRow = startRow + n + 2

    if(method == "CL"):
        writeRow(ws, infoRow, startCol, ["f"] + triangleData.get("f", []))
        ws.cell(row = infoRow + 2, column = startCol).value = "IBNR"
        ws.cell(row = infoRow + 2, column = startCol + 1).value = triangleData.get("IBNR")
    elif(method == "BF"):
        writeRow(ws, infoRow, startCol, ["Premium Sum"] + triangleData.get("premiums", triangleData.get("premiums", [])))
        writeRow(ws, infoRow + 1, startCol, ["ELR"] + triangleData.get("ELR", []))
        writeRow(ws, infoRow + 2, startCol, ["z"] + triangleData.get("z", []))
        ws.cell(row = infoRow + 4, column = startCol).value = "IBNR"
        ws.cell(row = infoRow + 4, column = startCol + 1).value = triangleData.get("IBNR")
    elif(method == "GLM"):
        writeRow(ws, infoRow, startCol, ["α"] + triangleData.get("alpha", []))
        writeRow(ws, infoRow + 1, startCol, ["β"] + triangleData.get("beta", []))
        writeRow(ws, infoRow + 2, startCol, ["γ"] + triangleData.get("gamma", []))
        writeRow(ws, infoRow + 3, startCol, ["Deviance", triangleData.get("deviance")])
        writeRow(ws, infoRow + 4, startCol, ["DF", triangleData.get("df")])
        ws.cell(row = infoRow + 6, column = startCol).value = "IBNR"
        ws.cell(row = infoRow + 6, column = startCol + 1).value = triangleData.get("IBNR")

    try:
        wb.save(filePath)
        messagebox.showinfo("Export", "IBNR results exported successfully!")
    except Exception as e:
        messagebox.showerror("Error", "Could not save Excel file!")
        print(f"{e}")
        return None


def createIBNRSection(root):
    createLabel(root, "IBNR Calculations", 40, 30, font = labelFont2)
    dataSourceVar = tk.StringVar(value = "Triangle")
    createLabel(root, "Data soure:", 40, 85)
    createRadioButton(root, 170, 80, "Triangle", dataSourceVar, "Triangle")
    createRadioButton(root, 170, 130, "Data file", dataSourceVar, "Data file")

    methodVar = tk.StringVar(value = "CL")
    createRadioButton(root, 40, 200, "Chain Ladder", methodVar, "CL")
    createRadioButton(root, 40, 250, "Bornhuetter-Ferguson", methodVar, "BF")
    createRadioButton(root, 40, 300, "GLM", methodVar, "GLM")
    methodFrame = tk.Frame(root, bg="#eeeeee")
    methodFrame.place(x = 20, y = 340, width = 270, height = 230)

    createLabel(root, "File:", 320, 80)
    createButton(root, text = "Browse", x = 380, y = 65, width = 140, height = 50,
        command = lambda: selectFile(root, selectedFileLabel))
    selectedFileLabel = createLabel(root, "", 320, 120)
    
    triangleInputFrame = tk.Frame(root, bg = "#eeeeee")
    triangleInputFrame.place(x = 750, y = 60, width = 380, height = 110)
    createLabel(triangleInputFrame, "Top left cell:", 0, 10)
    topLeftEntry = createEntry(triangleInputFrame, 150, 5)
    createLabel(root, "Period count:", 560, 70)
    periodCountEntry = createEntry(root, 700, 65, 70)    
    
    def updateMethodInputs(*args):
        if methodVar.get() == "CL":
            showCLInputs(methodFrame)
        elif methodVar.get() == "BF":
            showBFInputs(methodFrame)
        elif methodVar.get() == "GLM":
            showGLMInputs(methodFrame)
    methodVar.trace_add("write", updateMethodInputs)
    updateMethodInputs()

    dataFileInputFrame = tk.Frame(root, bg = "#eeeeee")
    dataFileInputFrame.place(x = 550, y = 60, width = 380, height = 150)
    periodLengthVar = tk.StringVar(value = "Year")
    createLabel(dataFileInputFrame, "Period length:", 50, 5)
    createRadioButton(dataFileInputFrame, 180, 0, "Month", periodLengthVar, "Month")
    createRadioButton(dataFileInputFrame, 180, 40, "Quarter", periodLengthVar, "Quarter")
    createRadioButton(dataFileInputFrame, 180, 80, "Year", periodLengthVar, "Year")
    valuationDateFrame = tk.Frame(root, bg = "#eeeeee")
    valuationDateFrame.place(x = 250, y = 700, width = 480, height = 150)
    dayEntry, monthEntry, yearEntry = createValuationDate(valuationDateFrame, x = 0, y = 0, dx = 150)
    dataFileInputFrame.place_forget()
    valuationDateFrame.place_forget()
    
    def updateDataSourceInputs(*args):
        if dataSourceVar.get() == "Triangle":
            dataFileInputFrame.place_forget()
            valuationDateFrame.place_forget()
            triangleInputFrame.place(x = 850, y = 60, width = 380, height = 110)
        else:
            triangleInputFrame.place_forget()
            dataFileInputFrame.place(x = 780, y = 60, width = 380, height=  150)
            valuationDateFrame.place(x = 250, y = 700, width = 480, height = 150)
    dataSourceVar.trace_add("write", updateDataSourceInputs)
    updateDataSourceInputs()

    triangleData = {"triangle": []}
    triangleFrame = tk.Frame(root, bg = "#eeeeee")
    triangleFrame.place(x= 320, y = 170, width = 900, height = 390)
    createTrianglePreview(triangleFrame)
    dataTypeVar = tk.StringVar(value = "Incremental")
    createLabel(root, "Data type:", 40, 580)
    dataTypeLabel = createLabel(root, dataTypeVar.get(), 160, 580)

    resultVar = tk.StringVar(value = "")
    resultLabel = tk.Label(root, textvariable = resultVar, font = ("Arial", 16, "bold"), bg = "#eeeeee")
    resultLabel.place(x = 900, y = 580)
    GLMInfoVar = tk.StringVar(value = "")
    GLMInfoLabel = tk.Label(root, textvariable = GLMInfoVar, font = ("Arial", 14), bg = "#eeeeee")
    GLMInfoLabel.place(x = 670, y = 580)
    def switchDataState():
        switchIBNRTriangleType(dataTypeVar, dataTypeLabel, triangleFrame, triangleData)

    createButton(root, text = "Switch data type", x = 40, y = 620, width = 200, height = 55, 
                 command = switchDataState)
    createButton(root, text = "Calculate", x = 40, y = 700, width = 160, height = 60, 
                 command = lambda: calculateIBNR(methodVar, dataTypeVar, dataTypeLabel, triangleFrame,
                                                 methodFrame, triangleData, resultVar, GLMInfoVar))
    createButton(root, text = "Export", x = 1000, y = 700, width = 160, height = 60,
                                command = lambda: exportIBNR(triangleData))
    createButton(root, text = "Import data", x = 320, y = 580, width = 160, height = 50,
                   command = lambda: importIBNRTriangleData(root, dataSourceVar, periodLengthVar, dataTypeVar, 
                                                            topLeftEntry, periodCountEntry,dayEntry, monthEntry,
                                                            yearEntry ,dataTypeLabel, triangleFrame, triangleData))


#RBNS calculation functions
def RBNS(root, dayEntry, monthEntry, yearEntry, resultVar, RBNS1Var, RBNS2Var, RBNS3Var):
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select source file first!")
        return None
    filePath = root.selectedInputFile
    valuationDate = getValuationDate(dayEntry, monthEntry, yearEntry)
    if valuationDate is None:
        return None
    wb, ws, mapping = getPolicyMapping(filePath, claimColnames, claimObligatoryColnames)
    if mapping is None:
        return None

    RBNS1 = 0
    RBNS2 = 0
    RBNS3 = 0
    skippedRows = 0
    for row in range(2, ws.max_row + 1):
        accidentDate = parseDate(ws.cell(row = row, column = mapping["accident date"]).value)
        reportDate = parseDate(ws.cell(row = row, column = mapping["report date"]).value)
        claimStatus = ws.cell(row = row, column = mapping["claim status"]).value
        caseReserve = ws.cell(row = row, column = mapping["case reserve"]).value
        if(accidentDate is None) or (reportDate is None) or (claimStatus is None) or (caseReserve is None):
            skippedRows += 1
            continue
        try:
            caseReserve = float(caseReserve)
        except Exception:
            messagebox.showerror("Error", f"Case Reserve should be a number!\nInvalid row: {row}")
            return None
        status = str(claimStatus).strip().lower()
        if(accidentDate > valuationDate) or (reportDate > valuationDate) or (status in ["settled", "rejected"]):
            continue

        if(status == "reported"):
            RBNS1 += caseReserve
        elif(status == "not settled"):
            paymentCount = 0
            paymentAmount = 0
            if("upcoming payment count" in mapping) and ("upcoming payment amount" in mapping):
                paymentCount = ws.cell(row = row, column = mapping["upcoming payment count"]).value
                paymentAmount = ws.cell(row = row, column = mapping["upcoming payment amount"]).value
                if(paymentCount not in [None, ""]) and (paymentAmount not in [None, ""]):
                    try:
                        paymentCount = float(paymentCount)
                        paymentAmount = float(paymentAmount)
                    except Exception:
                        messagebox.showerror("Error", f"Periodic payment data is invalid! Row number: {row}")
                        return None
                    RBNS3 += paymentCount*paymentAmount
            RBNS2 += caseReserve
        else:
            skippedRows += 1
            continue
    RBNS2 -= RBNS3
    RBNS = RBNS1 + RBNS2 + RBNS3
    if skippedRows > 0:
        messagebox.showwarning("Warning", f"{skippedRows} rows were skipped!")
    resultVar.set(f"RBNS = {round(RBNS, 2)}")
    RBNS1Var.set(f"RBNS1 = {round(RBNS1, 2)} ({round(RBNS1/RBNS*100, 2)} %)")
    RBNS2Var.set(f"RBNS2 = {round(RBNS2, 2)} ({round(RBNS2/RBNS*100, 2)} %)")
    RBNS3Var.set(f"RBNS3 = {round(RBNS3, 2)} ({round(RBNS3/RBNS*100, 2)} %)")
    
    


def createRBNSSection(root):
    createLabel(root, "RBNS Calculations", 40, 30, font = labelFont2)
    createLabel(root, "Select File:", 40, 100)
    selectedFileLabel = createLabel(root, "", 40, 160)
    createButton(root, text = "Browse", x = 200, y = 100, width = 140, height = 40, 
                 command = lambda: selectFile(root, selectedFileLabel))
    dayEntry, monthEntry, yearEntry = createValuationDate(root, x = 600, y = 100, dx = 200)
    
    createLabel(root, "RBNS components:", 40, 260, font = labelFont2)
    createLabel(root, "RBNS1: Reported claims", 40, 320)
    RBNS1Var = createResult(root, xPos = 400, yPos = 320)
    createLabel(root, "RBNS2: Not settled claims", 40, 370)
    RBNS2Var = createResult(root, xPos = 400, yPos = 370)
    createLabel(root, "RBNS3: Periodic payments", 40, 420)
    RBNS3Var = createResult(root, xPos = 400, yPos = 420)
    
    resultVar = tk.StringVar(value="")
    resultLabel = tk.Label(root, textvariable=resultVar, font=("Arial", 16, "bold"),
                           bg="#eeeeee", justify="left")
    resultLabel.place(x=40, y=560)
    resultVar = createResult(root, xPos = 40, yPos = 560)
    createButton(root, text = "Calculate", x = 40, y = 480, width = 180, height = 60,
                 command = lambda: RBNS(root, dayEntry, monthEntry, yearEntry, resultVar, RBNS1Var, RBNS2Var, RBNS3Var))


def createUPRSection(root):
    createLabel(root, "UPR Calculations", 40, 30, font=labelFont2)
    createLabel(root, "Select File:", 40, 100)
    selectedFileLabel = createLabel(root, "", 40, 160)
    createButton(root, text = "Browse", x = 200, y = 100, width = 140, height = 40,
                 command = lambda: selectFile(root, selectedFileLabel))

    dayEntry, monthEntry, yearEntry = createValuationDate(root)

    exportVar = tk.BooleanVar(value = False)
    exportCheckbox = tk.Checkbutton(root, text = "Export UPR to source file", variable = exportVar, font=("Arial", 15))
    exportCheckbox.place(x = 40, y = 250)
    createLabel(root, "Output column:", 40, 300)
    exportColumnEntry = createEntry(root, 200, 300, width = 120, height = 32)
    def updateExportState():
        if exportVar.get():
            exportColumnEntry.config(state = "normal")
        else:
            exportColumnEntry.delete(0, tk.END)
            exportColumnEntry.config(state = "disabled")
    exportVar.trace_add("write", lambda *args: updateExportState())
    updateExportState()
    
    resultVar = createResult(root, xPos = 40, yPos = 520)
    createButton(root, text = "Calculate", x = 40, y = 450, width = 180, height = 60,
                 command = lambda: UPR(root, dayEntry, monthEntry, yearEntry, exportVar, exportColumnEntry, resultVar))
    
    

def showReserveSection(frame, name):
    clearFrame(frame)
    if name == "IBNR":
        createIBNRSection(frame)
    elif name == "RBNS":
        createRBNSSection(frame)
    elif name == "UPR":
        createUPRSection(frame)


def createTab5(root):
    reserveBox, reserveTypeVar, filler = createCombobox(root, "Reserve type:", 30, 50, "str", "IBNR", ["IBNR", "RBNS", "UPR"])
    contentFrame = tk.Frame(root, bg="#eeeeee")
    contentFrame.place(x = 10, y = 80, width = 1250, height = 800)
    
    def onReserveChange(event = None):
        showReserveSection(contentFrame, reserveTypeVar.get())
    reserveBox.bind("<<ComboboxSelected>>", onReserveChange)
    showReserveSection(contentFrame, reserveTypeVar.get())




'''
DATA FILES
'''
def getSelectedFile(root):
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select source file first!")
        return None
    return root.selectedInputFile
    
    
def openDataFile(root):
    if not hasattr(root, "selectedInputFile"):
        messagebox.showerror("Error", "Please select source file first!")
        return None
    try:
        os.startfile(root.selectedInputFile)
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file:{root.selectedInputFile}")
        print(f"{e}")


def renameDataFile(root, selectedFileLabel):
    oldPath = getSelectedFile(root)
    oldDir = os.path.dirname(oldPath)
    oldName = os.path.basename(oldPath)
    oldExt = os.path.splitext(oldName)[1]
    newName = simpledialog.askstring("Rename file", "Enter new file name:")
    if not newName:
        return None
    newName = newName.strip()
    namePart, extPart = os.path.splitext(newName)
    if extPart.lower() not in (".xlsx", ".xlsm", ".xls"):
        newName += oldExt
    newPath = os.path.join(oldDir, newName)
    if os.path.exists(newPath):
        messagebox.showerror("Error", "File with this name already exists!")
        return None
    try:
        os.rename(oldPath, newPath)
        root.selectedInputFile = newPath
        selectedFileLabel.config(text = "Input file:" + os.path.basename(newPath))
        messagebox.showinfo("Rename file", "File renamed successfully.")
    except Exception as e:
        messagebox.showerror("Error", f"Could not rename file {oldName+oldExt} to {newName}")
        print(f"{e}")


def duplicateDataFile(root, selectedFileLabel):
    oldPath = getSelectedFile(root)
    oldDir = os.path.dirname(oldPath)
    oldName, oldExt = os.path.splitext(os.path.basename(oldPath))
    newPath = os.path.join(oldDir, oldName + " copy" + oldExt)
    counter = 1
    while os.path.exists(newPath):
        newPath = os.path.join(oldDir, f"{oldName} copy({counter}){oldExt}")
        counter += 1
    try:
        shutil.copy2(oldPath, newPath)
        root.selectedInputFile = newPath
        selectedFileLabel.config(text = "Input file:" + os.path.basename(newPath))
        messagebox.showinfo("Duplicate file", f"File duplicated:\n{os.path.basename(newPath)}")
    except Exception as e:
        messagebox.showerror("Error", f"Could not duplicate file {oldName}!")
        print(f"{e}")
        

def deleteDataFile(root, selectedFileLabel, mappingVar):
    filePath =  getSelectedFile(root)
    filename = os.path.basename(filePath)
    confirm = messagebox.askyesno("Delete file", f"Delete file '{filename}'?")
    if not confirm:
        return None
    try:
        os.remove(filePath)
        del root.selectedInputFile
        selectedFileLabel.config(text = "")
        mappingVar.set("")
        messagebox.showinfo("Delete file", "File deleted successfully!")
    except Exception as e:
        messagebox.showerror("Error", f"Could not delete file {filename}!")
        print(f"{e}")


def getMappingType(dataTypeVar):
    if dataTypeVar.get() == "Policies":
        return policyColnames, policyObligatoryColnames
    return claimColnames, claimObligatoryColnames


def getColName(number):
    try:
        result = ""
        while(number > 0):
            letterCode = number%26 + 64
            number = number//26
            result += chr(letterCode)
        return result
    except Exception as e:
        print(f"{e}")
        return None


def validateDataFile(root, dataTypeVar):
    filePath =  getSelectedFile(root)
    colnames, obligatoryColnames = getMappingType(dataTypeVar)
    wb, ws, mapping = getPolicyMapping(filePath, colnames, obligatoryColnames, doValidate = False)
    
    missingColumns = "Missing obligatory columns:\n"
    doAlert = False
    for colname in obligatoryColnames:
        if colname not in mapping:
            missingColumns += f"{colname};\n"
            doAlert = True
    if doAlert:
        messagebox.showerror("Validation", missingColumns)
    else:
        messagebox.showinfo("Validation", "File is valid and ready for use!")



def showDataMapping(root, dataTypeVar, mappingVar):
    filePath = getSelectedFile(root)
    colnames, obligatoryColnames = getMappingType(dataTypeVar)
    wb, ws, mapping = getPolicyMapping(filePath, colnames, obligatoryColnames, doValidate = False)
    if mapping is None:
        return None

    lines = []
    for colname in colnames:
        if colname in mapping:
            lines.append(f"{colname} - {getColName(mapping[colname])}")
        else:
            lines.append(f"{colname} - (empty)")
    mappingVar.set("\n\n".join(lines))



def isColumnEmpty(ws, col):
    for row in range(1, ws.max_row+1):
        if ws.cell(row = row, column = col).value is not None:
            return False
    return True


def swapColumns(ws, col1, col2):
    temp = ""
    for row in range(1, ws.max_row+1):
        temp = ws.cell(row = row, column = col1).value
        ws.cell(row = row, column = col1).value = ws.cell(row = row, column = col2).value
        ws.cell(row = row, column = col2).value = temp


def changeDataMapping(root, dataTypeVar, mappingVar):
    filePath =  getSelectedFile(root)
    colnames, obligatoryColnames = getMappingType(dataTypeVar)
    try:
        wb = load_workbook(filePath)
        ws = wb.active
    except Exception as e:
        messagebox.showerror("Error", f"Could not open file {filePath}!")
        print(f"{e}")
        return None
    filler, filler2, mapping = getPolicyMapping(filePath, colnames, [])
    if mapping is None:
        mapping = {}

    fromCol = simpledialog.askstring("Change mapping", "Which column do you want to move? Example: C")
    if not fromCol:
        return None
    toCol = simpledialog.askstring("Change mapping", "Where do you want to move it? Example: A")
    if not toCol:
        return None
    fromCol = getColNumber(fromCol)
    toCol = getColNumber(toCol)
    if(fromCol is None) or (toCol is None):
        messagebox.showerror("Error", "Invalid column format!")
        return None
    if(fromCol == toCol):
        messagebox.showerror("Mapping Error", "Source and target columns are the same.")
        return None
    
    targetColname = ws.cell(row = 1, column = toCol).value
    try:
        targetColname = targetColname.strip().lower()
    except:
        pass
    print(targetColname, mapping)
    if targetColname in mapping:
        confirm = messagebox.askyesno("Overwrite mapped column", 
                f"Target column named '{targetColname}' is a mapped colunmn.\nAre you sure you want to overwrite it?")
        if not confirm:
            return None
    elif not isColumnEmpty(ws, toCol):
        confirm = messagebox.askyesno( "Overwrite column", "Target column is not empty.\nAre you sure you want to overwrite it?")
        if not confirm:
            return None
        
    try:
        swapColumns(ws, fromCol, toCol)
        wb.save(filePath)
        messagebox.showinfo("Change mapping", "Column moved successfully!")
        showDataMapping(root, dataTypeVar, mappingVar)
    except Exception as e:
        messagebox.showerror("Error", f"Could not change mapping for columns {fromCol} and {toCol}")
        print(f"{e}")



def createTab6(root):
    createLabel(root, "Data", 50, 25, font = labelFont2)
    dataTypeVar = tk.StringVar(value = "Policies")
    createRadioButton(root, 250, 20, "Policies", dataTypeVar, "Policies")
    createRadioButton(root, 430, 20, "Claims", dataTypeVar, "Claims")
    createLabel(root, "File:", 60, 90)
    selectedFileLabel = createLabel(root, "", 60, 140)
    createButton(root, text = "Browse", x = 140, y = 75, width = 140, height = 55,
                 command = lambda: selectFile(root, selectedFileLabel))
    mappingVar = tk.StringVar(value = "")
    mappingLabel = tk.Label(root, textvariable = mappingVar, font = ("Arial", 16), bg = "#eeeeee", 
                            justify = "left", anchor = "nw", relief = "solid", bd = 2)
    mappingLabel.place(x = 600, y = 120, width = 350, height = 450)
    createButton(root, text = "Open", x = 60, y = 180, width = 120, height = 40,
                 command = lambda: openDataFile(root))
    createButton( root, text = "Rename", x = 60, y = 250, width = 120, height = 40,
                 command = lambda: renameDataFile(root, selectedFileLabel))
    createButton(root, text = "Duplicate", x = 60, y = 330, width = 120, height = 40,
                 command = lambda: duplicateDataFile(root, selectedFileLabel))
    createButton(root, text = "Validate", x = 60, y = 410, width = 120, height = 40,
                 command = lambda: validateDataFile(root, dataTypeVar))
    createButton(root, text = "Delete", x = 60, y = 520, width = 120, height = 40,
                 command = lambda: deleteDataFile(root, selectedFileLabel, mappingVar))
    createButton(root, text = "Show Mapping", x = 340, y = 190, width = 200, height = 50,
                 command = lambda: showDataMapping(root, dataTypeVar, mappingVar))
    createButton(root, text = "Change Mapping", x = 340, y = 500, width = 200, height = 50,
                 command = lambda: changeDataMapping(root, dataTypeVar, mappingVar))

    def clearMappingOnTypeChange(*args):
        mappingVar.set("")
    dataTypeVar.trace_add("write", clearMappingOnTypeChange)




def main():
    root = tk.Tk()
    root.title("Reserve calculator")
    root.geometry("1300x950")
    root.configure(bg = "#dddddd")
    
    notebook = ttk.Notebook(root)
    notebook.pack(expand = True, fill = "both", padx = 0, pady = 0)
    tabs = []
    for tabName in tabNames:
        tab = tk.Frame(notebook, bg = "#eeeeee")      #Here notebook is parent => tab is actually a tab in notebook
        notebook.add(tab, text = tabName)
        tabs.append(tab)
    style = ttk.Style()
    style.configure("TNotebook.Tab", width = 25, font = ("Times New Roman", 14, "bold"))
    createTab1(tabs[0])
    createTab2(tabs[1])
    createTab3(tabs[2])
    createTab4(tabs[3])
    createTab5(tabs[4])
    createTab6(tabs[5])
    
    root.mainloop()


main()